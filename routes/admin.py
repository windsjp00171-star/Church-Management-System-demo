import logging
# 管理員後台路由
from flask import Blueprint, session, redirect, url_for, render_template, request, jsonify, Response
from db import supabase
from routes.audit import log_action
from routes.decorators import admin_required, super_admin_required, staff_required
import secrets
import uuid
import io
import hmac
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timezone, timedelta

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

_TW = timezone(timedelta(hours=8))

def _tw_to_utc(s):
    """datetime-local 表單值（台灣時間，無時區）→ UTC ISO 字串"""
    if not s:
        return None
    try:
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=_TW)
        return dt.astimezone(timezone.utc).isoformat()
    except Exception:
        return None


_NAME_LABEL_EXACT = '參加者姓名'
_NAME_LABEL_FUZZY = ('姓名', '名字', '名稱')


def _build_name_override_map(reg_ids: list) -> dict:
    """從 registration_answers 找每筆報名的參加者姓名。
    優先順序：
      1. label == '參加者姓名'（標準欄位名）
      2. label 含 姓名/名字/名稱（舊活動相容）
      3. sort_order 最小的 text 欄位（最後防線）
    """
    if not reg_ids:
        return {}
    ans_res = supabase.table('registration_answers')\
        .select('registration_id, answer, event_fields(label, field_type, sort_order)')\
        .in_('registration_id', reg_ids)\
        .execute()

    exact  = {}  # { rid: answer }
    fuzzy  = {}  # { rid: (sort_order, answer) }
    text_f = {}  # { rid: (sort_order, answer) }

    for ans in (ans_res.data or []):
        rid   = ans['registration_id']
        field = ans.get('event_fields') or {}
        label = field.get('label', '')
        ftype = field.get('field_type', '')
        order = field.get('sort_order', 9999)
        val   = (ans.get('answer') or '').strip()
        if not val:
            continue
        if label == _NAME_LABEL_EXACT:
            exact[rid] = val
        elif any(kw in label for kw in _NAME_LABEL_FUZZY):
            if rid not in fuzzy or order < fuzzy[rid][0]:
                fuzzy[rid] = (order, val)
        if ftype == 'text':
            if rid not in text_f or order < text_f[rid][0]:
                text_f[rid] = (order, val)

    result = {}
    for rid in set(list(exact) + list(fuzzy) + list(text_f)):
        if rid in exact:
            result[rid] = exact[rid]
        elif rid in fuzzy:
            result[rid] = fuzzy[rid][1]
        else:
            result[rid] = text_f[rid][1]
    return result


def _generate_csrf_token():
    """為當前 session 產生（或取出）CSRF token"""
    if '_csrf_token' not in session:
        session['_csrf_token'] = secrets.token_hex(24)
    return session['_csrf_token']


@admin_bp.before_request
def csrf_protect():
    """所有後台 POST 請求都需要帶 X-CSRF-Token 標頭"""
    if request.method in ('POST', 'PUT', 'DELETE', 'PATCH'):
        token = request.headers.get('X-CSRF-Token') or request.form.get('_csrf_token')
        expected = session.get('_csrf_token')
        if not expected or not token or not hmac.compare_digest(token, expected):
            return jsonify({'error': 'CSRF token 驗證失敗，請重新整理頁面'}), 403



def can_manage_event(event):
    """檢查當前登入管理員是否有權限管理（編輯/刪除）此活動。
    超級管理員 or 活動建立者 → 有權限。
    若活動沒有建立者記錄（舊資料）→ 允許所有管理員。
    """
    if session.get('is_super_admin'):
        return True
    created_by = event.get('created_by')
    if not created_by:
        return True
    return session.get('user_id') == created_by


# =====================
# 後台首頁
# =====================

@admin_bp.route('/')
@admin_required
def index():
    """後台首頁"""
    return render_template('admin/index.html',
        has_finance_access=_has_finance_access())


@admin_bp.route('/calendar')
@admin_required
def calendar_page():
    """月曆視圖 → 統一行事曆頁面"""
    return redirect('/calendar')


@admin_bp.route('/api/calendar')
@admin_required
def calendar_api():
    """月曆資料 API：指定 year/month 回傳活動+門訓堂次"""
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))

    # 該月份的起訖（UTC 字串）
    from calendar import monthrange
    _, last_day = monthrange(year, month)
    start = f"{year:04d}-{month:02d}-01T00:00:00+00:00"
    end   = f"{year:04d}-{month:02d}-{last_day:02d}T23:59:59+00:00"

    # 活動（依 event_start 過濾）
    events_result = supabase.table('events')\
        .select('id, title, event_start, is_open, location')\
        .gte('event_start', start).lte('event_start', end)\
        .order('event_start').execute()

    # 門訓堂次（依 scheduled_at 過濾）
    sessions_result = supabase.table('course_sessions')\
        .select('id, title, session_number, scheduled_at, course_id')\
        .gte('scheduled_at', start).lte('scheduled_at', end)\
        .order('scheduled_at').execute()

    # 補上課程名稱
    sessions = sessions_result.data or []
    if sessions:
        cids = list({s['course_id'] for s in sessions})
        courses = supabase.table('courses').select('id, title')\
            .in_('id', cids).execute().data or []
        course_map = {c['id']: c['title'] for c in courses}
        for s in sessions:
            s['course_title'] = course_map.get(s['course_id'], '')

    # 教會手動行事
    church_result = supabase.table('church_events')\
        .select('id, title, event_date, description, color')\
        .gte('event_date', f"{year:04d}-{month:02d}-01")\
        .lte('event_date', f"{year:04d}-{month:02d}-{last_day:02d}")\
        .order('event_date').execute()

    items = []
    for e in (church_result.data or []):
        items.append({
            'type': 'church',
            'id': e['id'],
            'title': e['title'],
            'date': e['event_date'],
            'time': '',
            'description': e.get('description') or '',
            'color': e.get('color') or '#7b1fa2',
        })
    for e in (events_result.data or []):
        items.append({
            'type': 'event',
            'id': e['id'],
            'title': e['title'],
            'date': e['event_start'][:10],
            'time': e['event_start'][11:16] if e['event_start'] else '',
            'location': e.get('location') or '',
            'is_open': e.get('is_open', False),
        })
    for s in sessions:
        items.append({
            'type': 'session',
            'id': s['id'],
            'course_id': s['course_id'],
            'title': f"{s['course_title']} 第{s['session_number']}堂",
            'subtitle': s['title'],
            'date': s['scheduled_at'][:10],
            'time': s['scheduled_at'][11:16] if s['scheduled_at'] else '',
        })

    return jsonify(items)


@admin_bp.route('/calendar/church/new', methods=['POST'])
@admin_required
def church_event_new():
    data = request.get_json() or {}
    title = (data.get('title') or '').strip()
    event_date = (data.get('event_date') or '').strip()
    if not title or not event_date:
        return jsonify({'error': '請填寫標題與日期'}), 400
    try:
        remind_days = int(data.get('remind_days') or 3)
    except (ValueError, TypeError):
        remind_days = 3
    result = supabase.table('church_events').insert({
        'title': title,
        'event_date': event_date,
        'end_date': data.get('end_date') or None,
        'description': (data.get('description') or '').strip() or None,
        'color': data.get('color') or '#7b1fa2',
        'created_by': session.get('user_id'),
        'remind_days': remind_days,
    }).execute()
    if data.get('notify_all') and result.data:
        new_id = result.data[0]['id']
        all_users = supabase.table('users').select('id').eq('member_type', 'member').execute().data or []
        from routes.notifications import batch_notify
        batch_notify(
            user_ids=[u['id'] for u in all_users],
            title=f'⛪ 教會行事 — {title}',
            body=f'活動日期：{event_date}',
            type='announcement',
            link='/calendar',
            ref_type='church_event',
            ref_id=new_id,
        )
    return jsonify({'success': True})


@admin_bp.route('/calendar/church/<event_id>/edit', methods=['POST'])
@admin_required
def church_event_edit(event_id):
    data = request.get_json() or {}
    title = (data.get('title') or '').strip()
    event_date = (data.get('event_date') or '').strip()
    if not title or not event_date:
        return jsonify({'error': '請填寫標題與日期'}), 400
    try:
        remind_days = int(data.get('remind_days') or 3)
    except (ValueError, TypeError):
        remind_days = 3
    supabase.table('church_events').update({
        'title': title,
        'event_date': event_date,
        'end_date': data.get('end_date') or None,
        'description': (data.get('description') or '').strip() or None,
        'color': data.get('color') or '#7b1fa2',
        'remind_days': remind_days,
    }).eq('id', event_id).execute()
    return jsonify({'success': True})


@admin_bp.route('/calendar/church/<event_id>/delete', methods=['POST'])
@admin_required
def church_event_delete(event_id):
    supabase.table('church_events').delete().eq('id', event_id).execute()
    return jsonify({'success': True})


# =====================
# 使用者管理
# =====================

@admin_bp.route('/users')
@admin_required
def users():
    """會員管理頁面（超級管理員專屬）"""
    if not session.get('is_super_admin'):
        return render_template('admin/forbidden.html'), 403
    return render_template('admin/users.html')


@admin_bp.route('/api/users/create', methods=['POST'])
@admin_required
def create_manual_user():
    """手動建立無 LINE 帳號的會員"""
    if not session.get('is_super_admin'):
        return jsonify({'error': '無權限'}), 403
    data = request.get_json() or {}
    name = (data.get('real_name') or '').strip()
    if not name:
        return jsonify({'error': '請輸入姓名'}), 400
    result = supabase.table('users').insert({
        'real_name': name,
        'display_name': name,
        'member_type': data.get('member_type', 'member'),
    }).execute()
    if not result.data:
        return jsonify({'error': '建立失敗'}), 500
    log_action('user.create', 'user', result.data[0]['id'], {'name': name})
    return jsonify({'success': True, 'id': result.data[0]['id']})


@admin_bp.route('/api/users/<user_id>/merge-line', methods=['POST'])
@admin_required
def merge_line_account(user_id):
    """將 LINE 帳號整合到手動建立的帳號：複製 LINE 資料、移轉關聯紀錄、刪除舊帳號"""
    if not session.get('is_super_admin'):
        return jsonify({'error': '無權限'}), 403
    data = request.get_json() or {}
    line_user_id = (data.get('line_user_id') or '').strip()
    if not line_user_id:
        return jsonify({'error': '請指定 LINE 帳號 ID'}), 400

    # 查目標帳號（手動帳號）
    target_res = supabase.table('users').select('*').eq('id', user_id).execute()
    if not target_res.data:
        return jsonify({'error': '找不到目標帳號'}), 404
    target = target_res.data[0]
    if target.get('line_user_id'):
        return jsonify({'error': '此帳號已綁定 LINE，無需整合'}), 400

    # 查 LINE 帳號
    line_res = supabase.table('users').select('*').eq('line_user_id', line_user_id).execute()
    if not line_res.data:
        return jsonify({'error': '找不到此 LINE 帳號'}), 404
    line_acct = line_res.data[0]
    old_id = line_acct['id']

    if old_id == user_id:
        return jsonify({'error': '不能與自己整合'}), 400

    # ① 先清除舊 LINE 帳號的 line_user_id（line_user_id 有 UNIQUE 約束，
    #    必須先釋放才能複製到手動帳號，否則會 UNIQUE 衝突）
    supabase.table('users').update({'line_user_id': None}).eq('id', old_id).execute()

    # ② 複製 LINE 資訊到手動帳號
    try:
        supabase.table('users').update({
            'line_user_id': line_user_id,
            'display_name': line_acct.get('display_name') or target.get('display_name'),
            'picture_url': line_acct.get('picture_url') or target.get('picture_url'),
        }).eq('id', user_id).execute()
    except Exception as e:
        # 還原舊帳號的 LINE 身份，避免卡在中間狀態
        supabase.table('users').update({'line_user_id': line_user_id}).eq('id', old_id).execute()
        return jsonify({'error': f'更新手動帳號失敗：{e}'}), 500

    # ③ 移轉所有關聯紀錄（清單與 schema.sql 同步維護）
    from routes.account_merge import transfer_user_records
    failed = transfer_user_records(old_id, user_id)

    # ④ 刪除舊 LINE 帳號
    try:
        supabase.table('users').delete().eq('id', old_id).execute()
    except Exception as e:
        return jsonify({'error': f'刪除舊帳號失敗（可能仍有未移轉的關聯資料）：{e}'}), 500

    log_action('user.merge_line_admin', 'user', user_id, {
        'merged_from': old_id, 'line_user_id': line_user_id,
        'failed_tables': failed or None,
    })
    return jsonify({'success': True})


@admin_bp.route('/api/users/search')
@admin_required
def search_users():
    """搜尋使用者 API（會員管理頁用）
    ?q=       關鍵字（LINE 暱稱或真實姓名）
    ?type=    member | visitor | all（預設 all）
    """
    if not session.get('is_admin'):
        return jsonify({'error': '無權限'}), 403

    keyword     = request.args.get('q', '').strip()
    member_type = request.args.get('type', 'all').strip()   # member / visitor / all
    fields = 'id, display_name, picture_url, is_admin, is_super_admin, is_pastor, is_staff, is_blocked, group_tags, real_name, created_at, member_type, line_user_id'

    def apply_type_filter(query):
        if member_type == 'member':
            # member_type = 'member' 或 null（舊資料預設為會友）
            return query.or_('member_type.eq.member,member_type.is.null')
        elif member_type == 'visitor':
            return query.eq('member_type', 'visitor')
        return query  # all

    if not keyword:
        q = supabase.table('users').select(fields).order('real_name').order('created_at', desc=True).limit(200)
        q = apply_type_filter(q)
        data = q.execute().data or []
    else:
        # 搜尋 LINE 暱稱或真實姓名
        q1 = supabase.table('users').select(fields).ilike('display_name', f'%{keyword}%').limit(30)
        q2 = supabase.table('users').select(fields).ilike('real_name',    f'%{keyword}%').limit(30)
        q1 = apply_type_filter(q1)
        q2 = apply_type_filter(q2)
        by_display = q1.execute().data or []
        by_real    = q2.execute().data or []
        seen, data = set(), []
        for u in by_display + by_real:
            if u['id'] not in seen:
                seen.add(u['id'])
                data.append(u)

    return jsonify(data)


@admin_bp.route('/api/users/<user_id>/group-tag', methods=['POST'])
@admin_required
def set_group_tag(user_id):
    """設定使用者的小組標籤（僅超級管理員可操作）"""
    if not session.get('is_super_admin'):
        return jsonify({'error': '只有超級管理員可設定小組標籤'}), 403
    data = request.get_json() or {}
    tags = [t for t in data.get('group_tags', []) if t]
    supabase.table('users').update({'group_tags': tags}).eq('id', user_id).execute()
    return jsonify({'success': True, 'group_tags': tags})


# =====================
# 白名單管理
# =====================

@admin_bp.route('/api/whitelist/<ref_type>/<ref_id>')
@admin_required
def list_whitelist(ref_type, ref_id):
    """查看白名單"""
    rows = supabase.table('registration_whitelist')\
        .select('id, user_id, created_at')\
        .eq('ref_type', ref_type).eq('ref_id', ref_id).execute().data or []
    if not rows:
        return jsonify([])
    uids = [r['user_id'] for r in rows]
    users = supabase.table('users').select('id, real_name, display_name, picture_url, group_tags')\
        .in_('id', uids).execute().data or []
    user_map = {u['id']: u for u in users}
    for r in rows:
        r['user'] = user_map.get(r['user_id'], {})
    return jsonify(rows)


@admin_bp.route('/api/whitelist/<ref_type>/<ref_id>/add', methods=['POST'])
@admin_required
def add_whitelist(ref_type, ref_id):
    """加入白名單"""
    data = request.get_json() or {}
    user_id = data.get('user_id')
    if not user_id:
        return jsonify({'error': '請指定用戶'}), 400
    try:
        supabase.table('registration_whitelist').insert({
            'ref_type': ref_type, 'ref_id': ref_id, 'user_id': user_id
        }).execute()
        return jsonify({'success': True})
    except Exception:
        return jsonify({'error': '此用戶已在名單中'}), 400


@admin_bp.route('/api/whitelist/<ref_type>/<ref_id>/<user_id>/remove', methods=['POST'])
@admin_required
def remove_whitelist(ref_type, ref_id, user_id):
    """移出白名單"""
    supabase.table('registration_whitelist')\
        .delete().eq('ref_type', ref_type).eq('ref_id', ref_id).eq('user_id', user_id).execute()
    return jsonify({'success': True})


@admin_bp.route('/api/users/search')
@admin_required
def search_users_for_whitelist():
    """通用用戶搜尋（姓名）供白名單加人用"""
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])
    by_real = supabase.table('users').select('id, real_name, display_name, picture_url, group_tags')\
        .ilike('real_name', f'%{q}%').limit(10).execute().data or []
    by_line = supabase.table('users').select('id, real_name, display_name, picture_url, group_tags')\
        .ilike('display_name', f'%{q}%').limit(10).execute().data or []
    seen, merged = set(), []
    for u in by_real + by_line:
        if u['id'] not in seen:
            seen.add(u['id'])
            merged.append(u)
    return jsonify(merged[:15])


# =====================
# 圖片上傳
# =====================

ALLOWED_EXTENSIONS = {'jpg', 'jpeg', 'png', 'gif', 'webp'}

@admin_bp.route('/api/upload-image', methods=['POST'])
@admin_required
def upload_image():
    """上傳活動海報到 Supabase Storage"""
    file = request.files.get('image')
    if not file or not file.filename:
        return jsonify({'error': '請選擇圖片'}), 400
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({'error': '請上傳 JPG / PNG / GIF / WebP 圖片'}), 400

    filename = f"posters/{uuid.uuid4()}.{ext}"
    file_bytes = file.read()
    try:
        supabase.storage.from_('event-posters').upload(
            filename, file_bytes, {'content-type': file.content_type or 'image/jpeg'}
        )
        url = supabase.storage.from_('event-posters').get_public_url(filename)
        return jsonify({'success': True, 'url': url})
    except Exception as e:
        return jsonify({'error': f'上傳失敗：{str(e)}'}), 500


# =====================
# 小組管理
# =====================

@admin_bp.route('/groups')
@admin_required
def groups_page():
    """小組管理頁面（僅超級管理員）"""
    if not session.get('is_super_admin'):
        return render_template('admin/forbidden.html'), 403
    return render_template('admin/groups.html')


@admin_bp.route('/api/groups')
@admin_required
def list_groups():
    """撈所有小組（供前端下拉選單使用）"""
    result = supabase.table('groups').select('*').order('sort_order').execute()
    return jsonify(result.data or [])


@admin_bp.route('/api/groups', methods=['POST'])
@admin_required
def add_group():
    """新增服事角色標籤（僅超級管理員）"""
    if not session.get('is_super_admin'):
        return jsonify({'error': '只有超級管理員可新增標籤'}), 403
    data = request.get_json() or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': '標籤名稱不能為空'}), 400
    try:
        # 取目前最大排序值
        existing = supabase.table('groups').select('sort_order').order('sort_order', desc=True).limit(1).execute()
        next_order = (existing.data[0]['sort_order'] + 1) if existing.data else 1
        result = supabase.table('groups').insert({
            'name': name, 'sort_order': next_order, 'is_primary': False
        }).execute()
        return jsonify({'success': True, 'group': result.data[0]})
    except Exception as e:
        return jsonify({'error': '標籤名稱已存在或發生錯誤'}), 400


@admin_bp.route('/api/groups/<group_id>/delete', methods=['POST'])
@admin_required
def delete_group(group_id):
    """刪除小組（僅超級管理員）"""
    if not session.get('is_super_admin'):
        return jsonify({'error': '只有超級管理員可刪除小組'}), 403
    supabase.table('groups').delete().eq('id', group_id).execute()
    return jsonify({'success': True})


@admin_bp.route('/api/users/<user_id>/toggle-admin', methods=['POST'])
@admin_required
def toggle_admin(user_id):
    """切換使用者的管理員權限（超級管理員限定）"""
    if not session.get('is_super_admin'):
        return jsonify({'error': '僅超級管理員可操作'}), 403
    if user_id == session.get('user_id'):
        return jsonify({'error': '不能修改自己的權限'}), 400

    result = supabase.table('users')\
        .select('is_admin, is_super_admin, display_name')\
        .eq('id', user_id)\
        .execute()

    if not result.data:
        return jsonify({'error': '找不到此用戶'}), 404

    current = result.data[0]['is_admin']
    new_value = not current
    update = {'is_admin': new_value}
    # 取消管理員時，同步取消超管
    if not new_value:
        update['is_super_admin'] = False
    supabase.table('users').update(update).eq('id', user_id).execute()
    log_action('user.set_admin' if new_value else 'user.remove_admin',
               'user', user_id, {'name': result.data[0]['display_name']})

    return jsonify({
        'success': True,
        'is_admin': new_value,
        'is_super_admin': False if not new_value else result.data[0]['is_super_admin'],
        'display_name': result.data[0]['display_name']
    })


@admin_bp.route('/api/users/<user_id>/toggle-super-admin', methods=['POST'])
@admin_required
def toggle_super_admin(user_id):
    """切換使用者的超級管理員權限（超級管理員限定）"""
    if not session.get('is_super_admin'):
        return jsonify({'error': '僅超級管理員可操作'}), 403
    if user_id == session.get('user_id'):
        return jsonify({'error': '不能修改自己的權限'}), 400

    result = supabase.table('users')\
        .select('is_admin, is_super_admin, display_name')\
        .eq('id', user_id)\
        .execute()

    if not result.data:
        return jsonify({'error': '找不到此用戶'}), 404

    current = result.data[0]['is_super_admin']
    new_value = not current
    update = {'is_super_admin': new_value}
    # 設為超管時同步開啟管理員
    if new_value:
        update['is_admin'] = True
    supabase.table('users').update(update).eq('id', user_id).execute()
    log_action('user.set_super_admin' if new_value else 'user.remove_super_admin',
               'user', user_id, {'name': result.data[0]['display_name']})

    return jsonify({
        'success': True,
        'is_admin': True if new_value else result.data[0]['is_admin'],
        'is_super_admin': new_value,
        'display_name': result.data[0]['display_name']
    })


@admin_bp.route('/api/users/<user_id>/toggle-block', methods=['POST'])
@admin_required
def toggle_block_user(user_id):
    """封鎖 / 解封用戶（超級管理員限定）"""
    if not session.get('is_super_admin'):
        return jsonify({'error': '僅超級管理員可操作'}), 403
    if user_id == session.get('user_id'):
        return jsonify({'error': '不能封鎖自己'}), 400

    result = supabase.table('users')\
        .select('is_blocked, display_name')\
        .eq('id', user_id).execute()
    if not result.data:
        return jsonify({'error': '找不到此用戶'}), 404

    current = result.data[0].get('is_blocked') or False
    new_value = not current
    supabase.table('users').update({'is_blocked': new_value}).eq('id', user_id).execute()
    log_action('user.block' if new_value else 'user.unblock',
               'user', user_id, {'name': result.data[0]['display_name']})
    return jsonify({
        'success': True,
        'is_blocked': new_value,
        'display_name': result.data[0]['display_name'],
    })


# =====================
# Demo 洽詢名單
# =====================

@admin_bp.route('/logs')
@admin_required
def audit_logs_page():
    """稽核日誌查詢頁（超級管理員限定）"""
    if not session.get('is_super_admin'):
        return render_template('admin/forbidden.html'), 403
    page = max(1, request.args.get('page', 1, type=int))
    per_page = 50
    action_filter = request.args.get('action', '').strip()

    query = supabase.table('audit_logs').select('*', count='exact')\
        .order('created_at', desc=True)
    if action_filter:
        query = query.eq('action', action_filter)
    try:
        result = query.range((page - 1) * per_page, page * per_page - 1).execute()
        logs, total = result.data or [], result.count or 0
    except Exception:
        logging.getLogger(__name__).warning('稽核日誌查詢失敗（資料表可能尚未建立）', exc_info=True)
        logs, total = [], 0
    total_pages = max(1, (total + per_page - 1) // per_page)

    from routes.audit import ACTION_LABELS
    return render_template('admin/audit_logs.html',
                           logs=logs, action_labels=ACTION_LABELS,
                           action_filter=action_filter, page=page,
                           total_pages=total_pages, total=total)


@admin_bp.route('/contact-leads')
@admin_required
def contact_leads():
    try:
        rows = supabase.table('contact_leads')\
            .select('*').order('submitted_at', desc=True).limit(200).execute().data or []
    except Exception:
        rows = []
    return render_template('admin/contact_leads.html', rows=rows)


# =====================
# 小組回報管理（cell_groups）
# =====================

@admin_bp.route('/cell-groups')
@admin_required
def cell_groups_page():
    if not session.get('is_super_admin'):
        return render_template('admin/forbidden.html'), 403
    return render_template('admin/cell_groups.html')


@admin_bp.route('/api/cell-groups')
@admin_required
def list_cell_groups():
    groups = supabase.table('cell_groups').select('*').order('name').execute().data or []
    for g in groups:
        leaders = supabase.table('cell_group_leaders')\
            .select('user_id, users(id, real_name, display_name)')\
            .eq('group_id', g['id']).execute().data or []
        g['leaders'] = [l['users'] for l in leaders if l.get('users')]
    return jsonify(groups)


@admin_bp.route('/api/cell-groups', methods=['POST'])
@admin_required
def create_cell_group():
    if not session.get('is_super_admin'):
        return jsonify({'error': '僅超級管理員可操作'}), 403
    data = request.get_json() or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': '名稱不能為空'}), 400
    gather_day = data.get('weekly_gather_day', 0)
    result = supabase.table('cell_groups').insert({
        'name': name, 'weekly_gather_day': gather_day, 'is_active': True
    }).execute()
    return jsonify({'success': True, 'group': result.data[0]})


@admin_bp.route('/api/cell-groups/<group_id>/leaders', methods=['POST'])
@admin_required
def set_cell_leader(group_id):
    if not session.get('is_super_admin'):
        return jsonify({'error': '僅超級管理員可操作'}), 403
    data = request.get_json() or {}
    user_id = data.get('user_id', '').strip()
    if not user_id:
        return jsonify({'error': '請指定用戶'}), 400
    existing = supabase.table('cell_group_leaders')\
        .select('id').eq('group_id', group_id).eq('user_id', user_id).execute().data
    if existing:
        return jsonify({'error': '此人已是該小組的小組長'}), 400
    supabase.table('cell_group_leaders').insert({'group_id': group_id, 'user_id': user_id}).execute()
    user = supabase.table('users').select('real_name, display_name').eq('id', user_id).execute().data
    return jsonify({'success': True, 'user': user[0] if user else {}})


@admin_bp.route('/api/cell-groups/<group_id>/leaders/<user_id>', methods=['DELETE'])
@admin_required
def remove_cell_leader(group_id, user_id):
    if not session.get('is_super_admin'):
        return jsonify({'error': '僅超級管理員可操作'}), 403
    supabase.table('cell_group_leaders').delete()\
        .eq('group_id', group_id).eq('user_id', user_id).execute()
    return jsonify({'success': True})


@admin_bp.route('/api/cell-groups/<group_id>/toggle-active', methods=['POST'])
@admin_required
def toggle_cell_group_active(group_id):
    if not session.get('is_super_admin'):
        return jsonify({'error': '僅超級管理員可操作'}), 403
    current = supabase.table('cell_groups').select('is_active').eq('id', group_id).execute().data
    if not current:
        return jsonify({'error': '找不到小組'}), 404
    new_val = not current[0]['is_active']
    supabase.table('cell_groups').update({'is_active': new_val}).eq('id', group_id).execute()
    return jsonify({'success': True, 'is_active': new_val})


@admin_bp.route('/api/cell-groups/<group_id>/members')
@admin_required
def list_cell_group_members(group_id):
    """取得小組的活躍成員清單（含 user_id 連結狀態與確認狀態）"""
    members = supabase.table('cell_members')\
        .select('id, name, user_id, is_confirmed')\
        .eq('group_id', group_id)\
        .eq('is_active', True)\
        .order('id').execute().data or []
    return jsonify(members)


@admin_bp.route('/api/cell-groups/<group_id>/members/<member_id>/confirm', methods=['POST'])
@admin_required
def confirm_cell_member(group_id, member_id):
    """確認（核准）自選小組申請"""
    supabase.table('cell_members').update({'is_confirmed': True})\
        .eq('id', member_id).eq('group_id', group_id).execute()
    return jsonify({'success': True})


@admin_bp.route('/api/cell-groups/<group_id>/members/<member_id>/reject', methods=['POST'])
@admin_required
def reject_cell_member(group_id, member_id):
    """拒絕（軟刪除）自選小組申請"""
    supabase.table('cell_members').update({'is_active': False})\
        .eq('id', member_id).eq('group_id', group_id).execute()
    return jsonify({'success': True})


@admin_bp.route('/api/users/members')
@admin_required
def list_members_simple():
    """供小組長選人用的簡易會員清單"""
    result = supabase.table('users').select('id, real_name, display_name').order('real_name').execute()
    return jsonify(result.data or [])


@admin_bp.route('/api/users/search-for-cell')
@admin_required
def search_users_for_cell():
    """搜尋使用者供連結牧養小組成員帳號用
    ?q=  關鍵字（真實姓名或 LINE 暱稱），回傳 [{id, name}] 最多 20 筆
    """
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])
    by_real = supabase.table('users').select('id, real_name, display_name')\
        .ilike('real_name', f'%{q}%').limit(20).execute().data or []
    by_line = supabase.table('users').select('id, real_name, display_name')\
        .ilike('display_name', f'%{q}%').limit(20).execute().data or []
    seen, merged = set(), []
    for u in by_real + by_line:
        if u['id'] not in seen:
            seen.add(u['id'])
            merged.append({
                'id': u['id'],
                'name': u.get('real_name') or u.get('display_name') or '—',
            })
    return jsonify(merged[:20])


@admin_bp.route('/api/users/<user_id>/toggle-pastor', methods=['POST'])
@admin_required
def toggle_pastor(user_id):
    if not session.get('is_super_admin'):
        return jsonify({'error': '僅超級管理員可操作'}), 403
    result = supabase.table('users').select('is_pastor, display_name').eq('id', user_id).execute()
    if not result.data:
        return jsonify({'error': '找不到此用戶'}), 404
    new_value = not bool(result.data[0].get('is_pastor', False))
    supabase.table('users').update({'is_pastor': new_value}).eq('id', user_id).execute()
    log_action('user.set_pastor' if new_value else 'user.remove_pastor',
               'user', user_id, {'name': result.data[0]['display_name']})
    return jsonify({'success': True, 'is_pastor': new_value, 'display_name': result.data[0]['display_name']})


@admin_bp.route('/api/users/<user_id>/toggle-staff', methods=['POST'])
@admin_required
def toggle_staff(user_id):
    if not session.get('is_super_admin'):
        return jsonify({'error': '僅超級管理員可操作'}), 403
    result = supabase.table('users').select('is_staff, display_name').eq('id', user_id).execute()
    if not result.data:
        return jsonify({'error': '找不到此用戶'}), 404
    new_value = not bool(result.data[0].get('is_staff', False))
    supabase.table('users').update({'is_staff': new_value}).eq('id', user_id).execute()
    log_action('user.set_staff' if new_value else 'user.remove_staff',
               'user', user_id, {'name': result.data[0]['display_name']})
    return jsonify({'success': True, 'is_staff': new_value, 'display_name': result.data[0]['display_name']})


# =====================
# 活動管理
# =====================

@admin_bp.route('/events')
@admin_required
def events():
    """活動列表"""
    result = supabase.table('events')\
        .select('*')\
        .order('created_at', desc=True)\
        .execute()
    events_data = result.data or []

    # 撈出所有建立者的 group_tags
    creator_ids = list({e['created_by'] for e in events_data if e.get('created_by')})
    group_tags_map = {}  # user_id → [tag, ...]
    if creator_ids:
        users_result = supabase.table('users')\
            .select('id, group_tags')\
            .in_('id', creator_ids)\
            .execute()
        for u in (users_result.data or []):
            group_tags_map[u['id']] = u.get('group_tags') or []

    # 把 group_tags 附加到每筆活動，並計算實際開放狀態
    now = datetime.now(timezone.utc).isoformat()
    for e in events_data:
        e['group_tags'] = group_tags_map.get(e.get('created_by'), [])
        # effective_open：手動開關 AND 在時間窗口內
        if not e.get('is_open'):
            e['effective_open'] = False
            e['open_reason'] = 'manual'          # 手動關閉
        elif e.get('reg_end') and e['reg_end'] < now:
            e['effective_open'] = False
            e['open_reason'] = 'expired'         # 時間自動截止
        elif e.get('reg_start') and e['reg_start'] > now:
            e['effective_open'] = False
            e['open_reason'] = 'not_yet'         # 尚未開始
        else:
            e['effective_open'] = True
            e['open_reason'] = 'open'

    return render_template('admin/events.html',
        events=events_data,
        current_user_id=session.get('user_id'),
        is_super_admin=session.get('is_super_admin', False)
    )


@admin_bp.route('/events/new', methods=['GET', 'POST'])
@admin_required
def event_new():
    """建立新活動"""
    if request.method == 'POST':
        data = request.get_json()

        if data.get('checkin_mode') == 'open' and (data.get('whitelist_enabled') or data.get('waitlist_enabled')):
            return jsonify({'error': '開放型簽到不能與「限定報名名單」或「候補名單」同時啟用'}), 400

        # 建立活動主體
        checkin_enabled = data.get('checkin_enabled', False)
        event_data = {
            'title': data.get('title'),
            'description': data.get('description'),
            'location': data.get('location'),
            'event_start': _tw_to_utc(data.get('event_start')),
            'event_end':   _tw_to_utc(data.get('event_end')),
            'reg_start':   _tw_to_utc(data.get('reg_start')),
            'reg_end':     _tw_to_utc(data.get('reg_end')),
            'capacity': int(data['capacity']) if data.get('capacity') else None,
            'fee': int(data.get('fee', 0)),
            'is_open': data.get('is_open', True),
            'created_by': session.get('user_id'),
            'checkin_enabled': checkin_enabled,
            'checkin_mode': data.get('checkin_mode', 'registered_only'),
            'allow_open_checkin': data.get('allow_open_checkin', False),
            'checkin_token': secrets.token_urlsafe(16) if checkin_enabled else None,
            'allow_multiple': data.get('allow_multiple', False),
            'allow_external_reg': data.get('allow_external_reg', False),
            'payment_enabled': data.get('payment_enabled', False),
            'party_animation': data.get('party_animation', False),
            'poster_url': data.get('poster_url') or None,
            'whitelist_enabled': data.get('whitelist_enabled', False),
            'waitlist_enabled': data.get('waitlist_enabled', False),
            'waitlist_deadline': data.get('waitlist_deadline') or None,
            'meal_options': data.get('meal_options') or None,
            'reminder_days': int(data.get('reminder_days') or 3),
        }
        try:
            event_result = supabase.table('events').insert(event_data).execute()
        except Exception as e:
            return jsonify({'error': f'建立失敗：{str(e)}'}), 500
        event_id = event_result.data[0]['id']

        # 建立自訂欄位
        try:
            fields = data.get('fields', [])
            for i, field in enumerate(fields):
                if field.get('label'):
                    supabase.table('event_fields').insert({
                        'event_id': event_id,
                        'label': field['label'],
                        'field_type': field.get('field_type', 'text'),
                        'options': field.get('options', ''),
                        'is_required': field.get('is_required', False),
                        'sort_order': i,
                        'condition_json': field.get('condition_json') or None,
                    }).execute()
        except Exception as e:
            return jsonify({'error': f'欄位建立失敗：{str(e)}'}), 500

        return jsonify({'success': True, 'event_id': event_id})

    return render_template('admin/event_form.html', event=None, fields=[])


@admin_bp.route('/events/<event_id>/edit', methods=['GET', 'POST'])
@admin_required
def event_edit(event_id):
    """編輯活動"""
    # 先撈活動確認權限
    _ev = supabase.table('events').select('*').eq('id', event_id).execute()
    if not _ev.data:
        return '找不到此活動', 404
    if not can_manage_event(_ev.data[0]):
        return render_template('admin/forbidden.html'), 403

    if request.method == 'POST':
        data = request.get_json()

        if data.get('checkin_mode') == 'open' and (data.get('whitelist_enabled') or data.get('waitlist_enabled')):
            return jsonify({'error': '開放型簽到不能與「限定報名名單」或「候補名單」同時啟用'}), 400

        checkin_enabled = data.get('checkin_enabled', False)

        # 編輯活動：若已有 token 則保留（不讓 QR Code 失效），否則新產生
        existing_token = None
        if checkin_enabled:
            existing_event = supabase.table('events').select('checkin_token').eq('id', event_id).execute()
            if existing_event.data:
                existing_token = existing_event.data[0].get('checkin_token')
            if not existing_token:
                existing_token = secrets.token_urlsafe(16)

        event_data = {
            'title': data.get('title'),
            'description': data.get('description'),
            'location': data.get('location'),
            'event_start': _tw_to_utc(data.get('event_start')),
            'event_end':   _tw_to_utc(data.get('event_end')),
            'reg_start':   _tw_to_utc(data.get('reg_start')),
            'reg_end':     _tw_to_utc(data.get('reg_end')),
            'capacity': int(data['capacity']) if data.get('capacity') else None,
            'fee': int(data.get('fee', 0)),
            'is_open': data.get('is_open', True),
            'checkin_enabled': checkin_enabled,
            'checkin_mode': data.get('checkin_mode', 'registered_only'),
            'allow_open_checkin': data.get('allow_open_checkin', False),
            'checkin_token': existing_token,
            'allow_multiple': data.get('allow_multiple', False),
            'allow_external_reg': data.get('allow_external_reg', False),
            'payment_enabled': data.get('payment_enabled', False),
            'party_animation': data.get('party_animation', False),
            'poster_url': data.get('poster_url') or None,
            'whitelist_enabled': data.get('whitelist_enabled', False),
            'waitlist_enabled': data.get('waitlist_enabled', False),
            'waitlist_deadline': data.get('waitlist_deadline') or None,
            'meal_options': data.get('meal_options') or None,
            'reminder_days': int(data.get('reminder_days') or 3),
        }
        try:
            supabase.table('events').update(event_data).eq('id', event_id).execute()
        except Exception as e:
            return jsonify({'error': f'更新失敗：{str(e)}'}), 500

        # 更新欄位：永不 hard delete，移除的欄位改為 is_archived=True
        # 這樣既有報名答案永遠不會消失，名單頁仍可顯示歷史答案
        try:
            fields = data.get('fields', [])
            submitted_ids = set()

            for i, field in enumerate(fields):
                if not field.get('label'):
                    continue
                fid = field.get('id') or None
                field_payload = {
                    'event_id': event_id,
                    'label': field['label'],
                    'field_type': field.get('field_type', 'text'),
                    'options': field.get('options', ''),
                    'is_required': field.get('is_required', False),
                    'sort_order': i,
                    'condition_json': field.get('condition_json') or None,
                    'is_archived': False,  # 出現在提交中 = 使用中
                }
                if fid:
                    supabase.table('event_fields').update(field_payload)\
                        .eq('id', fid).eq('event_id', event_id).execute()
                    submitted_ids.add(fid)
                else:
                    result = supabase.table('event_fields').insert(field_payload).execute()
                    if result.data:
                        submitted_ids.add(result.data[0]['id'])

            # 未出現在提交中的欄位 → 軟刪除（is_archived=True），保留答案
            existing = supabase.table('event_fields').select('id')\
                .eq('event_id', event_id).execute()
            for ef in (existing.data or []):
                if ef['id'] not in submitted_ids:
                    supabase.table('event_fields').update({'is_archived': True})\
                        .eq('id', ef['id']).execute()

        except Exception as e:
            return jsonify({'error': f'欄位更新失敗：{str(e)}'}), 500

        return jsonify({'success': True})

    event = _ev.data[0]
    # 編輯表單只顯示未封存的欄位（封存欄位不再讓管理員看到/編輯）
    fields = supabase.table('event_fields').select('*')\
        .eq('event_id', event_id).eq('is_archived', False).order('sort_order').execute().data
    return render_template('admin/event_form.html', event=event, fields=fields)


@admin_bp.route('/events/<event_id>/clone', methods=['POST'])
@admin_required
def event_clone(event_id):
    """複製活動（複製基本設定與自訂欄位，日期清空，預設關閉報名）"""
    event_result = supabase.table('events').select('*').eq('id', event_id).execute()
    if not event_result.data:
        return jsonify({'error': '找不到此活動'}), 404
    src = event_result.data[0]

    # 建立新活動（日期清空、預設關閉報名）
    new_event = supabase.table('events').insert({
        'title': src['title'] + '（複製）',
        'description': src.get('description'),
        'location': src.get('location'),
        'capacity': src.get('capacity'),
        'fee': src.get('fee', 0),
        'is_open': False,
        'checkin_enabled': False,
        'checkin_mode': src.get('checkin_mode', 'registered_only'),
        'created_by': session.get('user_id'),
    }).execute()
    new_id = new_event.data[0]['id']

    # 複製自訂欄位
    fields_result = supabase.table('event_fields')\
        .select('*').eq('event_id', event_id).order('sort_order').execute()
    for f in (fields_result.data or []):
        supabase.table('event_fields').insert({
            'event_id': new_id,
            'label': f['label'],
            'field_type': f['field_type'],
            'options': f.get('options', ''),
            'is_required': f.get('is_required', False),
            'sort_order': f['sort_order'],
            'condition_json': f.get('condition_json'),
        }).execute()

    return jsonify({'success': True, 'new_id': new_id})


@admin_bp.route('/events/<event_id>/checkin-live')
@staff_required
def checkin_live(event_id):
    """即時簽到狀況頁面"""
    event_result = supabase.table('events').select('*').eq('id', event_id).execute()
    if not event_result.data:
        return '找不到此活動', 404
    return render_template('admin/checkin_live.html', event=event_result.data[0])


@admin_bp.route('/events/<event_id>/checkin-display')
@admin_required
def checkin_display(event_id):
    """現場大螢幕展示頁：有人簽到就播派對動畫"""
    ev = supabase.table('events').select('*').eq('id', event_id).execute()
    if not ev.data:
        return '找不到此活動', 404
    if not can_manage_event(ev.data[0]):
        return render_template('admin/forbidden.html'), 403
    return render_template('admin/checkin_display.html', event=ev.data[0])


@admin_bp.route('/events/<event_id>/checkin-live/data')
@staff_required
def checkin_live_data(event_id):
    """即時簽到狀況 JSON API（前端輪詢用）"""
    # 撈所有已報名（registered + walk_in）紀錄
    regs_result = supabase.table('registrations')\
        .select('*')\
        .eq('event_id', event_id)\
        .in_('status', ['registered', 'walk_in'])\
        .order('created_at')\
        .execute()
    regs = regs_result.data or []

    # 撈使用者資料（含小組標籤）；外部報名者 user_id 為 null，跳過
    user_map = {}
    if regs:
        user_ids = list({r['user_id'] for r in regs if r.get('user_id')})
        if user_ids:
            users_result = supabase.table('users')\
                .select('id, display_name, real_name, picture_url, group_tags')\
                .in_('id', user_ids)\
                .execute()
            for u in (users_result.data or []):
                user_map[u['id']] = u

    # 從報名答案找真實姓名（label 含姓名/名字 → 第一個 text 欄位）
    reg_ids = [r['id'] for r in regs]
    name_override_map = _build_name_override_map(reg_ids)

    # 組成回傳資料
    taipei_tz = timezone(timedelta(hours=8))
    checked = []
    waiting = []
    for reg in regs:
        user = user_map.get(reg['user_id'], {})
        override_name = name_override_map.get(reg['id'])
        display_name = (
            override_name
            or user.get('real_name') or user.get('display_name')
            or reg.get('guest_name') or '外部來賓'
        )
        # 代報時 override_name 是參加者姓名，與帳號持有者不同，小組標籤不屬於參加者
        user_own_names = {user.get('real_name'), user.get('display_name')} - {None, ''}
        is_proxy = bool(override_name and override_name not in user_own_names)
        item = {
            'id': reg['id'],
            'display_name': display_name,
            'picture_url': user.get('picture_url', ''),
            'group_tags': [] if is_proxy else (user.get('group_tags') or []),
            'status': reg.get('status', ''),
            'checked_in': reg.get('checked_in', False),
            'created_at': '',
        }
        if reg.get('created_at'):
            try:
                t = datetime.fromisoformat(reg['created_at'].replace('Z', '+00:00'))
                item['created_at'] = t.astimezone(taipei_tz).strftime('%H:%M')
            except Exception:
                item['created_at'] = reg['created_at'][11:16]

        if reg.get('checked_in'):
            checked.append(item)
        else:
            waiting.append(item)

    return jsonify({
        'total': len(regs),
        'checked_count': len(checked),
        'waiting_count': len(waiting),
        'checked': checked,
        'waiting': waiting,
    })


@admin_bp.route('/events/<event_id>/checkin-live/export')
@admin_required
def checkin_live_export(event_id):
    """匯出簽到名單 Excel"""
    event_result = supabase.table('events').select('*').eq('id', event_id).execute()
    if not event_result.data:
        return '找不到此活動', 404
    event = event_result.data[0]

    # 撈所有已報名（registered + walk_in）紀錄
    regs = supabase.table('registrations')\
        .select('*')\
        .eq('event_id', event_id)\
        .in_('status', ['registered', 'walk_in'])\
        .order('checked_in', desc=True)\
        .execute().data or []

    # 撈使用者資料
    user_map = {}
    if regs:
        user_ids = list({r['user_id'] for r in regs if r.get('user_id')})
        if user_ids:
            users = supabase.table('users')\
                .select('id, display_name, real_name, picture_url, group_tags')\
                .in_('id', user_ids)\
                .execute().data or []
            for u in users:
                user_map[u['id']] = u

    # 從報名答案找真實姓名（label 含姓名/名字 → 第一個 text 欄位）
    reg_ids_all = [r['id'] for r in regs]
    name_override_map = _build_name_override_map(reg_ids_all)

    taipei_tz = timezone(timedelta(hours=8))

    # 建立 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = '簽到名單'

    # 標題列樣式
    header_font  = Font(bold=True, color='FFFFFF')
    header_fill  = PatternFill('solid', fgColor='06C755')
    center_align = Alignment(horizontal='center', vertical='center')

    headers = ['姓名', '小組', '狀態', '簽到時間', '備註']
    col_widths = [18, 20, 10, 16, 12]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = center_align
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22

    for reg in regs:
        user  = user_map.get(reg['user_id'], {})
        name  = (
            name_override_map.get(reg['id'])
            or user.get('real_name') or user.get('display_name')
            or reg.get('guest_name') or '外部來賓'
        )
        tags  = user.get('group_tags') or []
        group = '、'.join(tags) if isinstance(tags, list) else str(tags)

        status_label = '✅ 已簽到' if reg.get('checked_in') else '⏳ 未到'
        note = '現場到場' if reg.get('status') == 'walk_in' else ''

        # 簽到時間（registrations 可能沒有 checked_in_at，先用 updated_at 或空）
        checkin_time = ''
        raw_time = reg.get('checked_in_at') or (reg.get('updated_at') if reg.get('checked_in') else '')
        if raw_time:
            try:
                t = datetime.fromisoformat(raw_time.replace('Z', '+00:00'))
                checkin_time = t.astimezone(taipei_tz).strftime('%m/%d %H:%M')
            except Exception:
                checkin_time = raw_time[11:16]

        ws.append([name, group, status_label, checkin_time, note])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_title = event.get('title', '簽到名單').replace('/', '_')
    filename = f"{safe_title}_簽到名單.xlsx"
    return Response(
        output.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename*=UTF-8\'\'{quote(filename)}'}
    )


@admin_bp.route('/events/<event_id>/toggle', methods=['POST'])
@admin_required
def event_toggle(event_id):
    """手動開關報名"""
    result = supabase.table('events').select('*').eq('id', event_id).execute()
    if not result.data:
        return jsonify({'error': '找不到此活動'}), 404
    event = result.data[0]
    if not can_manage_event(event):
        return jsonify({'error': '你沒有權限修改此活動'}), 403
    current = event['is_open']
    supabase.table('events').update({'is_open': not current}).eq('id', event_id).execute()
    return jsonify({'success': True, 'is_open': not current})


@admin_bp.route('/events/<event_id>/delete', methods=['POST'])
@admin_required
def event_delete(event_id):
    """刪除活動（連同報名紀錄、自訂欄位一併清除）"""
    result = supabase.table('events').select('*').eq('id', event_id).execute()
    if not result.data:
        return jsonify({'error': '找不到此活動'}), 404
    if not can_manage_event(result.data[0]):
        return jsonify({'error': '你沒有權限刪除此活動'}), 403

    try:
        # 1. 找出所有報名 ID
        reg_rows = supabase.table('registrations')\
            .select('id').eq('event_id', event_id).execute().data or []
        reg_ids = [r['id'] for r in reg_rows]

        # 2. 刪 registration_answers（FK → registrations）
        if reg_ids:
            supabase.table('registration_answers')\
                .delete().in_('registration_id', reg_ids).execute()

        # 3. 刪 registrations（FK → events）
        supabase.table('registrations').delete().eq('event_id', event_id).execute()

        # 4. 刪 event_fields（FK → events）
        supabase.table('event_fields').delete().eq('event_id', event_id).execute()

        # 5. 刪 event 本體
        supabase.table('events').delete().eq('id', event_id).execute()

    except Exception as e:
        return jsonify({'error': f'刪除失敗：{str(e)}'}), 500

    log_action('event.delete', 'event', event_id,
               {'title': result.data[0].get('title')})
    return jsonify({'success': True})


# =====================
# 報名名單管理
# =====================

@admin_bp.route('/events/<event_id>/registrations')
@staff_required
def registrations(event_id):
    """查看活動報名名單"""
    # 撈活動資料
    event_result = supabase.table('events').select('*').eq('id', event_id).execute()
    if not event_result.data:
        return '找不到此活動', 404
    event = event_result.data[0]

    # 撈所有報名紀錄（只撈 registered 狀態）
    regs_result = supabase.table('registrations')\
        .select('*')\
        .eq('event_id', event_id)\
        .eq('status', 'registered')\
        .order('created_at')\
        .execute()
    regs = regs_result.data or []

    # 撈使用者資料（只撈有 user_id 的 LINE 用戶）
    user_map = {}
    if regs:
        user_ids = list({r['user_id'] for r in regs if r.get('user_id')})
        if user_ids:
            users_result = supabase.table('users')\
                .select('id, real_name, display_name, picture_url')\
                .in_('id', user_ids)\
                .execute()
            for u in (users_result.data or []):
                user_map[u['id']] = u

    # 撈自訂欄位定義（含已封存，才能顯示歷史答案）
    fields_result = supabase.table('event_fields')\
        .select('*')\
        .eq('event_id', event_id)\
        .order('sort_order')\
        .execute()
    fields = fields_result.data or []

    # 撈所有報名的自訂欄位答案，整理成 { registration_id: { field_id: answer } }
    answer_map = {}
    name_override_map = {}
    if regs:
        reg_ids = [r['id'] for r in regs]
        answers_result = supabase.table('registration_answers')\
            .select('registration_id, field_id, answer')\
            .in_('registration_id', reg_ids)\
            .execute()
        for ans in (answers_result.data or []):
            rid = ans['registration_id']
            if rid not in answer_map:
                answer_map[rid] = {}
            answer_map[rid][ans['field_id']] = ans['answer']
        name_override_map = _build_name_override_map(reg_ids)

    return render_template('admin/registrations.html',
        event=event,
        regs=regs,
        user_map=user_map,
        fields=fields,
        answer_map=answer_map,
        name_override_map=name_override_map,
        is_admin=session.get('is_admin', False) or session.get('is_super_admin', False),
    )


# =====================
# QR Code 簽到管理
# =====================

@admin_bp.route('/events/<event_id>/reg-qrcode')
@admin_required
def event_reg_qrcode(event_id):
    """顯示活動報名頁的 QR Code（掃碼直接進報名）"""
    event_result = supabase.table('events').select('*').eq('id', event_id).execute()
    if not event_result.data:
        return '找不到此活動', 404
    event = event_result.data[0]
    reg_url = request.host_url.rstrip('/') + f"/event/{event_id}"
    return render_template('admin/reg_qrcode.html', event=event, reg_url=reg_url)


@admin_bp.route('/events/<event_id>/qrcode')
@staff_required
def event_qrcode(event_id):
    """顯示活動的電子簽到 QR Code"""
    event_result = supabase.table('events').select('*').eq('id', event_id).execute()
    if not event_result.data:
        return '找不到此活動', 404
    event = event_result.data[0]

    if not event.get('checkin_enabled') or not event.get('checkin_token'):
        return '此活動尚未開啟電子簽到', 400

    # 組出完整的簽到 URL（根據目前 request 的 host）
    checkin_url = request.host_url.rstrip('/') + f"/checkin/{event_id}/{event['checkin_token']}"
    return render_template('admin/qrcode.html', event=event, checkin_url=checkin_url)


@admin_bp.route('/events/<event_id>/registrations/<reg_id>/checkin', methods=['POST'])
@staff_required
def toggle_checkin(event_id, reg_id):
    """切換報到狀態"""
    result = supabase.table('registrations')\
        .select('checked_in')\
        .eq('id', reg_id)\
        .eq('event_id', event_id)\
        .execute()

    if not result.data:
        return jsonify({'error': '找不到此報名紀錄'}), 404

    current = result.data[0]['checked_in']
    new_value = not current
    update_data = {'checked_in': new_value}
    if new_value:
        # 簽到時記錄時間
        update_data['checked_in_at'] = datetime.now(timezone.utc).isoformat()
    else:
        # 取消簽到時清除時間戳
        update_data['checked_in_at'] = None
    supabase.table('registrations').update(update_data).eq('id', reg_id).execute()
    return jsonify({'success': True, 'checked_in': new_value})


# =====================
# 代簽到搜尋
# =====================

@admin_bp.route('/events/<event_id>/checkin-live/search')
@staff_required
def checkin_search(event_id):
    """從全部會員搜尋（供同工代簽，含純簽到活動）"""
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])

    # 從所有會員中搜尋姓名（分兩次查詢避免 filter 字串注入）
    pattern = f'%{q}%'
    r1 = supabase.table('users').select('id, real_name, display_name, picture_url')\
        .ilike('real_name', pattern).limit(20).execute().data or []
    r2 = supabase.table('users').select('id, real_name, display_name, picture_url')\
        .ilike('display_name', pattern).limit(20).execute().data or []
    seen = set()
    users_result = []
    for u in r1 + r2:
        if u['id'] not in seen:
            seen.add(u['id'])
            users_result.append(u)

    if not users_result:
        return jsonify([])

    # 查這些人在此活動的報名狀態
    user_ids = [u['id'] for u in users_result]
    regs = supabase.table('registrations')\
        .select('id, checked_in, user_id, status')\
        .eq('event_id', event_id)\
        .in_('user_id', user_ids)\
        .in_('status', ['registered', 'walk_in'])\
        .execute().data or []
    reg_map = {r['user_id']: r for r in regs}

    results = []
    for u in users_result:
        name = u.get('real_name') or u.get('display_name') or ''
        reg = reg_map.get(u['id'])
        results.append({
            'user_id': u['id'],
            'reg_id': reg['id'] if reg else None,
            'name': name,
            'picture_url': u.get('picture_url', ''),
            'checked_in': reg['checked_in'] if reg else False,
            'status': reg['status'] if reg else 'none',  # none = 尚無紀錄
        })

    return jsonify(results)


@admin_bp.route('/events/<event_id>/proxy-checkin', methods=['POST'])
@staff_required
def proxy_checkin(event_id):
    """同工代替會友簽到（含無報名紀錄的純簽到情境）"""
    body = request.get_json() or {}
    user_id = body.get('user_id')
    reg_id = body.get('reg_id')

    if not user_id and not reg_id:
        return jsonify({'error': '缺少識別資訊'}), 400

    # 撈活動設定，用於 checkin_mode 判斷
    event_result = supabase.table('events')\
        .select('checkin_mode').eq('id', event_id).execute()
    if not event_result.data:
        return jsonify({'error': '找不到此活動'}), 404
    checkin_mode = event_result.data[0].get('checkin_mode', 'registered_only')

    now_utc = datetime.now(timezone.utc).isoformat()

    if reg_id:
        # 已有報名紀錄（含外部報名者）→ 確認狀態合法再簽到
        reg_result = supabase.table('registrations')\
            .select('id, status, checked_in')\
            .eq('id', reg_id)\
            .eq('event_id', event_id)\
            .execute().data or []
        if not reg_result:
            return jsonify({'error': '找不到報名紀錄'}), 404
        reg = reg_result[0]
        if reg.get('checked_in'):
            return jsonify({'error': '此人已完成簽到'}), 400
        if checkin_mode == 'registered_only' and reg.get('status') != 'registered':
            return jsonify({'error': '此報名狀態無法簽到（僅限已報名者）'}), 400
        supabase.table('registrations')\
            .update({'checked_in': True, 'checked_in_at': now_utc})\
            .eq('id', reg_id).execute()
    else:
        # 無報名紀錄 → 僅限開放模式才能建立 walk_in
        if checkin_mode == 'registered_only':
            return jsonify({'error': '此活動僅限已報名者簽到，請先為此人完成報名'}), 400
        existing = supabase.table('registrations')\
            .select('id, checked_in')\
            .eq('event_id', event_id)\
            .eq('user_id', user_id)\
            .eq('status', 'walk_in')\
            .execute().data or []
        if existing:
            if existing[0].get('checked_in'):
                return jsonify({'error': '此人已完成簽到'}), 400
            supabase.table('registrations')\
                .update({'checked_in': True, 'checked_in_at': now_utc})\
                .eq('id', existing[0]['id']).execute()
        else:
            supabase.table('registrations').insert({
                'event_id': event_id,
                'user_id': user_id,
                'status': 'walk_in',
                'checked_in': True,
                'checked_in_at': now_utc,
                'source': 'proxy',
            }).execute()

    return jsonify({'success': True})


# =====================
# 付款狀態管理
# =====================

@admin_bp.route('/events/<event_id>/registrations/export')
@admin_required
def export_registrations(event_id):
    """匯出報名名單為 Excel（.xlsx）"""
    # 撈活動資料
    event_result = supabase.table('events').select('*').eq('id', event_id).execute()
    if not event_result.data:
        return '找不到此活動', 404
    event = event_result.data[0]

    # 撈報名紀錄（registered + walk_in）
    regs_result = supabase.table('registrations')\
        .select('*')\
        .eq('event_id', event_id)\
        .in_('status', ['registered', 'walk_in'])\
        .order('created_at')\
        .execute()
    regs = regs_result.data or []

    # 撈使用者資料（過濾掉外部報名者的 null user_id）
    user_map = {}
    if regs:
        user_ids = list({r['user_id'] for r in regs if r.get('user_id')})
        if user_ids:
            users_result = supabase.table('users')\
                .select('id, display_name')\
                .in_('id', user_ids)\
                .execute()
            for u in (users_result.data or []):
                user_map[u['id']] = u

    # 撈自訂欄位定義（含封存，確保 Excel 匯出也能看到歷史答案）
    fields_result = supabase.table('event_fields')\
        .select('*')\
        .eq('event_id', event_id)\
        .order('sort_order')\
        .execute()
    fields = fields_result.data or []

    # 撈所有自訂欄位答案
    answer_map = {}
    if regs:
        reg_ids = [r['id'] for r in regs]
        answers_result = supabase.table('registration_answers')\
            .select('registration_id, field_id, answer')\
            .in_('registration_id', reg_ids)\
            .execute()
        for ans in (answers_result.data or []):
            rid = ans['registration_id']
            if rid not in answer_map:
                answer_map[rid] = {}
            answer_map[rid][ans['field_id']] = ans['answer']

    # ===== 建立 Excel 檔案 =====
    wb = Workbook()
    ws = wb.active
    ws.title = '報名名單'

    # 標題列樣式
    header_fill = PatternFill(start_color='06C755', end_color='06C755', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    center = Alignment(horizontal='center', vertical='center')

    # 建立標題列
    headers = ['序號', '姓名（LINE）', '報名類型', '報名時間', '付款狀態', '報到狀態']
    for field in fields:
        headers.append(field['label'])

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # 付款/報到狀態中文對照
    payment_map = {'paid': '已付款', 'waived': '免收費', 'unpaid': '未付款'}
    status_map = {'registered': '一般報名', 'walk_in': '現場到場'}

    # 寫入每筆報名資料
    taipei_tz = timezone(timedelta(hours=8))
    for row_idx, reg in enumerate(regs, 2):
        user = user_map.get(reg['user_id'], {})
        answers = answer_map.get(reg['id'], {})

        # 報名時間轉台灣時間
        reg_time = ''
        if reg.get('created_at'):
            try:
                t = datetime.fromisoformat(reg['created_at'].replace('Z', '+00:00'))
                reg_time = t.astimezone(taipei_tz).strftime('%Y/%m/%d %H:%M')
            except Exception:
                reg_time = reg['created_at'][:16]

        row_data = [
            row_idx - 1,
            user.get('display_name') or reg.get('guest_name') or '外部來賓',
            status_map.get(reg.get('status', ''), reg.get('status', '')),
            reg_time,
            payment_map.get(reg.get('payment_status', 'unpaid'), '未付款'),
            '已報到' if reg.get('checked_in') else '未報到',
        ]
        # 自訂欄位答案
        for field in fields:
            row_data.append(answers.get(field['id'], ''))

        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=value)

    # 自動調整欄寬
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max(10, min(max_len + 4, 40))

    # 輸出為下載檔案
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # 檔名（含活動名稱）
    safe_title = event['title'].replace('/', '-').replace('\\', '-')
    filename = f"報名名單_{safe_title}.xlsx"

    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename*=UTF-8\'\'{quote(filename)}'}
    )


@admin_bp.route('/events/<event_id>/registrations/<reg_id>/cancel', methods=['POST'])
@admin_required
def cancel_registration(event_id, reg_id):
    """後台刪除（取消）報名紀錄"""
    result = supabase.table('registrations')\
        .select('id').eq('id', reg_id).eq('event_id', event_id).execute()
    if not result.data:
        return jsonify({'error': '找不到此報名紀錄'}), 404
    supabase.table('registrations').update({'status': 'cancelled'})\
        .eq('id', reg_id).execute()
    return jsonify({'success': True})


@admin_bp.route('/events/<event_id>/registrations/<reg_id>/edit', methods=['POST'])
@admin_required
def edit_registration(event_id, reg_id):
    """後台：編輯報名資料（外部匯入補填 or 一般報名修正）"""
    data = request.get_json() or {}

    # 確認記錄存在
    reg = supabase.table('registrations')\
        .select('id, user_id').eq('id', reg_id).eq('event_id', event_id).execute()
    if not reg.data:
        return jsonify({'error': '找不到此報名紀錄'}), 404

    is_guest = not reg.data[0].get('user_id')

    # 外部報名才能改名字和電話
    if is_guest:
        update_data = {}
        if data.get('guest_name', '').strip():
            update_data['guest_name'] = data['guest_name'].strip()
        if 'guest_phone' in data:
            update_data['guest_phone'] = data['guest_phone'].strip() or None
        if update_data:
            supabase.table('registrations').update(update_data).eq('id', reg_id).execute()

    # 補填 / 更新自訂欄位答案（先刪再寫，確保乾淨）
    answers = data.get('answers', [])  # [{field_id, answer}]
    if answers:
        supabase.table('registration_answers').delete().eq('registration_id', reg_id).execute()
        for ans in answers:
            if ans.get('field_id') and str(ans.get('answer', '')).strip():
                supabase.table('registration_answers').insert({
                    'registration_id': reg_id,
                    'field_id': ans['field_id'],
                    'answer': str(ans['answer']).strip(),
                }).execute()

    return jsonify({'success': True})


@admin_bp.route('/events/<event_id>/kiosk')
@admin_required
def kiosk_checkin(event_id):
    """Kiosk 報到頁：觸控友善，列出所有報名者，點名簽到（外部匯入用）"""
    event_result = supabase.table('events').select('*').eq('id', event_id).execute()
    if not event_result.data:
        return '找不到此活動', 404
    event = event_result.data[0]

    regs_result = supabase.table('registrations')\
        .select('*')\
        .eq('event_id', event_id)\
        .eq('status', 'registered')\
        .order('created_at')\
        .execute()
    regs = regs_result.data or []

    # 撈 LINE 用戶顯示名稱
    user_map = {}
    line_user_ids = [r['user_id'] for r in regs if r.get('user_id')]
    if line_user_ids:
        users_result = supabase.table('users')\
            .select('id, real_name, display_name, picture_url')\
            .in_('id', line_user_ids)\
            .execute()
        for u in (users_result.data or []):
            user_map[u['id']] = u

    # 從報名答案找真實姓名（label 含姓名/名字 → 第一個 text 欄位）
    reg_ids = [r['id'] for r in regs]
    name_override_map = _build_name_override_map(reg_ids)

    # 整理顯示名稱：報名表姓名答案 > guest_name（外部）> 用戶真名
    for r in regs:
        if r.get('user_id') and r['user_id'] in user_map:
            u = user_map[r['user_id']]
            r['_display_name'] = (
                name_override_map.get(r['id'])
                or u.get('real_name') or u.get('display_name') or '未知'
            )
            r['_picture'] = u.get('picture_url') or ''
            r['_source'] = 'line'
        else:
            r['_display_name'] = r.get('guest_name') or '未知來賓'
            r['_picture'] = ''
            r['_source'] = r.get('source') or 'import'

    # 排序：未報到優先，再按姓名
    regs.sort(key=lambda r: (r.get('checked_in', False), r['_display_name']))

    return render_template('admin/kiosk.html', event=event, regs=regs)


@admin_bp.route('/events/<event_id>/registrations/add-guest', methods=['POST'])
@admin_required
def add_guest_registration(event_id):
    """管理員手動新增外部（非LINE）報名者"""
    data = request.get_json() or {}
    guest_name = (data.get('guest_name') or '').strip()
    guest_phone = (data.get('guest_phone') or '').strip()
    if not guest_name:
        return jsonify({'error': '請輸入姓名'}), 400

    now = datetime.now(timezone(timedelta(hours=8))).isoformat()
    result = supabase.table('registrations').insert({
        'event_id': event_id,
        'user_id': None,
        'guest_name': guest_name,
        'guest_phone': guest_phone or None,
        'status': 'registered',
        'checked_in': False,
        'source': 'import',
        'created_at': now,
    }).execute()
    if result.data:
        return jsonify({'success': True, 'id': result.data[0]['id']})
    return jsonify({'error': '新增失敗'}), 500


@admin_bp.route('/events/<event_id>/registrations/<reg_id>/payment', methods=['POST'])
@admin_required
def toggle_payment(event_id, reg_id):
    """切換付款狀態：unpaid → paid → waived → unpaid"""
    result = supabase.table('registrations')\
        .select('payment_status')\
        .eq('id', reg_id)\
        .eq('event_id', event_id)\
        .execute()

    if not result.data:
        return jsonify({'error': '找不到此報名紀錄'}), 404

    current = result.data[0].get('payment_status', 'unpaid')
    # 三態循環：未付款 → 已付款 → 免收費 → 未付款
    cycle = {'unpaid': 'paid', 'paid': 'waived', 'waived': 'unpaid'}
    new_status = cycle.get(current, 'paid')

    supabase.table('registrations').update({'payment_status': new_status}).eq('id', reg_id).execute()
    return jsonify({'success': True, 'payment_status': new_status})


# ══════════════════════════════════════════
# 後台：每日經文管理
# ══════════════════════════════════════════

@admin_bp.route('/verses')
@admin_required
def admin_verses():
    from datetime import date
    from routes.event import VERSE_THEMES
    import time
    verses = supabase.table('daily_verses').select('*').order('sort_order').execute().data or []
    try:
        custom_themes = supabase.table('verse_custom_themes').select('*').order('sort_order').execute().data or []
    except Exception:
        custom_themes = []
    return render_template('admin/admin_verses.html', verses=verses,
                           themes=VERSE_THEMES,
                           custom_themes=custom_themes,
                           now=int(time.time()),
                           today=date.today().strftime('%Y/%m/%d'))


@admin_bp.route('/verses/upload-watermark', methods=['POST'])
@admin_required
def upload_verse_watermark():
    import os, imghdr
    f = request.files.get('image')
    if not f or not f.filename:
        return jsonify({'success': False, 'error': '請選擇圖片'})
    data = f.read()
    if imghdr.what(None, data) not in ('png', 'jpeg', 'gif', 'webp'):
        return jsonify({'success': False, 'error': '只接受 PNG / JPG / GIF / WebP'})
    dest = os.path.join(os.path.dirname(__file__), '..', 'static', 'img', 'lion.png')
    with open(dest, 'wb') as out:
        out.write(data)
    return jsonify({'success': True})


@admin_bp.route('/verses/upload-theme', methods=['POST'])
@admin_required
def upload_verse_theme():
    import imghdr, time as _time
    name = (request.form.get('name') or '').strip()
    symbol = (request.form.get('symbol') or '✝').strip() or '✝'
    text_mode = request.form.get('text_mode', 'light')
    if text_mode not in ('light', 'dark'):
        text_mode = 'light'
    f = request.files.get('image')
    if not name:
        return jsonify({'success': False, 'error': '請填寫樣式名稱'})
    if not f or not f.filename:
        return jsonify({'success': False, 'error': '請選擇圖片'})
    data = f.read()
    fmt = imghdr.what(None, data)
    if fmt not in ('png', 'jpeg', 'webp'):
        return jsonify({'success': False, 'error': '只接受 PNG / JPG / WebP'})
    ext = 'jpg' if fmt == 'jpeg' else fmt
    filename = f"theme_{int(_time.time() * 1000)}.{ext}"
    try:
        supabase.storage.from_('verse-themes').upload(
            path=filename, file=data,
            file_options={'content-type': f.content_type or f'image/{fmt}', 'upsert': 'true'}
        )
        url = supabase.storage.from_('verse-themes').get_public_url(filename)
    except Exception as e:
        return jsonify({'success': False, 'error': f'Storage 上傳失敗：{e}'})
    existing = supabase.table('verse_custom_themes').select('sort_order').order('sort_order', desc=True).limit(1).execute().data
    next_order = (existing[0]['sort_order'] + 1) if existing else 0
    row = supabase.table('verse_custom_themes').insert({
        'name': name, 'image_url': url,
        'symbol': symbol, 'text_mode': text_mode,
        'sort_order': next_order,
    }).execute().data[0]
    return jsonify({'success': True, 'theme': row})


@admin_bp.route('/verses/custom-theme/<theme_id>/delete', methods=['POST'])
@admin_required
def delete_verse_custom_theme(theme_id):
    row = supabase.table('verse_custom_themes').select('image_url').eq('id', theme_id).execute().data
    if row:
        try:
            filename = row[0]['image_url'].split('verse-themes/')[-1].split('?')[0]
            supabase.storage.from_('verse-themes').remove([filename])
        except Exception:
            logging.getLogger(__name__).warning('忽略非關鍵錯誤', exc_info=True)
    supabase.table('verse_custom_themes').delete().eq('id', theme_id).execute()
    return jsonify({'success': True})


@admin_bp.route('/verses/new', methods=['POST'])
@admin_required
def admin_verse_new():
    data = request.get_json() or {}
    text = (data.get('text') or '').strip()
    ref = (data.get('ref') or '').strip()
    if not text:
        return jsonify({'error': '請填寫經文內容'}), 400
    existing = supabase.table('daily_verses').select('sort_order')\
        .order('sort_order', desc=True).limit(1).execute().data
    next_order = (existing[0]['sort_order'] + 1) if existing else 1
    result = supabase.table('daily_verses').insert({
        'text': text, 'ref': ref or None,
        'sort_order': next_order, 'is_active': True,
    }).execute()
    return jsonify({'success': True, 'id': result.data[0]['id']})


@admin_bp.route('/verses/<verse_id>/edit', methods=['POST'])
@admin_required
def admin_verse_edit(verse_id):
    data = request.get_json() or {}
    text = (data.get('text') or '').strip()
    if not text:
        return jsonify({'error': '請填寫經文內容'}), 400
    supabase.table('daily_verses').update({
        'text': text,
        'ref': (data.get('ref') or '').strip() or None,
        'is_active': bool(data.get('is_active', True)),
    }).eq('id', verse_id).execute()
    return jsonify({'success': True})


@admin_bp.route('/verses/<verse_id>/delete', methods=['POST'])
@admin_required
def admin_verse_delete(verse_id):
    supabase.table('daily_verses').delete().eq('id', verse_id).execute()
    return jsonify({'success': True})


# ══════════════════════════════════════════
# 後台：快捷連結管理
# ══════════════════════════════════════════

@admin_bp.route('/portal-links')
@super_admin_required
def admin_portal_links():
    links = supabase.table('portal_links').select('*').order('sort_order').execute().data or []
    return render_template('admin/admin_links.html', links=links)


@admin_bp.route('/portal-links/new', methods=['POST'])
@super_admin_required
def admin_portal_link_new():
    data = request.get_json() or {}
    title = (data.get('title') or '').strip()
    url = (data.get('url') or '').strip()
    if not title or not url:
        return jsonify({'error': '請填寫標題與網址'}), 400
    existing = supabase.table('portal_links').select('sort_order')\
        .order('sort_order', desc=True).limit(1).execute().data
    next_order = (existing[0]['sort_order'] + 1) if existing else 1
    result = supabase.table('portal_links').insert({
        'title': title,
        'subtitle': (data.get('subtitle') or '').strip() or None,
        'url': url,
        'emoji': (data.get('emoji') or '🔗').strip(),
        'border_color': (data.get('border_color') or '#888888').strip(),
        'is_staff_only': bool(data.get('is_staff_only', False)),
        'member_only': bool(data.get('member_only', False)),
        'is_external': bool(data.get('is_external', True)),
        'sort_order': next_order,
        'is_active': True,
    }).execute()
    return jsonify({'success': True, 'id': result.data[0]['id']})


@admin_bp.route('/portal-links/<link_id>/edit', methods=['POST'])
@super_admin_required
def admin_portal_link_edit(link_id):
    data = request.get_json() or {}
    title = (data.get('title') or '').strip()
    url = (data.get('url') or '').strip()
    if not title or not url:
        return jsonify({'error': '請填寫標題與網址'}), 400
    supabase.table('portal_links').update({
        'title': title,
        'subtitle': (data.get('subtitle') or '').strip() or None,
        'url': url,
        'emoji': (data.get('emoji') or '🔗').strip(),
        'border_color': (data.get('border_color') or '#888888').strip(),
        'is_staff_only': bool(data.get('is_staff_only', False)),
        'member_only': bool(data.get('member_only', False)),
        'is_external': bool(data.get('is_external', True)),
        'is_active': bool(data.get('is_active', True)),
    }).eq('id', link_id).execute()
    return jsonify({'success': True})


@admin_bp.route('/portal-links/<link_id>/delete', methods=['POST'])
@super_admin_required
def admin_portal_link_delete(link_id):
    supabase.table('portal_links').delete().eq('id', link_id).execute()
    return jsonify({'success': True})


@admin_bp.route('/portal-links/reorder', methods=['POST'])
@super_admin_required
def reorder_portal_links():
    data = request.get_json() or {}
    order = data.get('order', [])
    for item in order:
        link_id = item.get('id')
        sort_order = item.get('sort_order')
        if link_id and sort_order is not None:
            supabase.table('portal_links').update({'sort_order': sort_order}).eq('id', link_id).execute()
    return jsonify({'success': True})


# ══════════════════════════════════════════
# API：小組 is_staff 切換
# ══════════════════════════════════════════

@admin_bp.route('/api/groups/<group_id>/toggle-staff', methods=['POST'])
@admin_required
def toggle_group_staff(group_id):
    """切換小組的同工標記（is_staff）"""
    current = supabase.table('groups').select('is_staff').eq('id', group_id).execute().data
    if not current:
        return jsonify({'error': '找不到小組'}), 404
    new_val = not bool(current[0].get('is_staff', False))
    supabase.table('groups').update({'is_staff': new_val}).eq('id', group_id).execute()
    return jsonify({'success': True, 'is_staff': new_val})


# =====================
# 週報管理
# =====================

@admin_bp.route('/bulletins')
@admin_required
def admin_bulletins():
    """週報管理頁"""
    result = supabase.table('weekly_bulletins')\
        .select('*')\
        .order('bulletin_date', desc=True)\
        .execute()
    bulletins = result.data or []
    return render_template('admin/bulletins.html', bulletins=bulletins)


@admin_bp.route('/api/upload-bulletin', methods=['POST'])
@admin_required
def upload_bulletin():
    """上傳週報 PDF"""
    file = request.files.get('pdf')
    if not file or not file.filename:
        return jsonify({'error': '請選擇 PDF 檔案'}), 400
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({'error': '請上傳 PDF 檔案'}), 400

    bulletin_date = request.form.get('bulletin_date')
    title = request.form.get('title', '').strip()
    if not bulletin_date:
        return jsonify({'error': '請填寫週報日期'}), 400
    if not title:
        title = f"{bulletin_date} 週報"

    filename = f"bulletins/{uuid.uuid4()}.pdf"
    file_bytes = file.read()
    try:
        # 嘗試建立 bucket（已存在時忽略錯誤）
        try:
            supabase.storage.create_bucket('bulletins', options={'public': True})
        except Exception:
            logging.getLogger(__name__).warning('忽略非關鍵錯誤', exc_info=True)
        supabase.storage.from_('bulletins').upload(
            filename, file_bytes, {'content-type': 'application/pdf'}
        )
        url = supabase.storage.from_('bulletins').get_public_url(filename)
    except Exception as e:
        return jsonify({'error': f'上傳失敗：{str(e)}'}), 500

    supabase.table('weekly_bulletins').insert({
        'title': title,
        'bulletin_date': bulletin_date,
        'pdf_url': url,
        'created_by': session.get('user_id'),
    }).execute()

    return jsonify({'success': True})


@admin_bp.route('/bulletins/<bulletin_id>/delete', methods=['POST'])
@admin_required
def admin_bulletin_delete(bulletin_id):
    """刪除週報"""
    supabase.table('weekly_bulletins').delete().eq('id', bulletin_id).execute()
    return jsonify({'success': True})


# =====================
# 小組討論管理
# =====================

import json as _json

@admin_bp.route('/group-discussions')
@admin_required
def group_discussions_page():
    """小組討論管理頁"""
    import json as _json
    items = supabase.table('group_discussions')\
        .select('*')\
        .order('created_at', desc=True)\
        .execute().data or []
    # questions 欄位可能是 JSON 字串，統一轉成 list
    for item in items:
        q = item.get('questions')
        if isinstance(q, str):
            try:
                item['questions'] = _json.loads(q)
            except Exception:
                item['questions'] = []
    return render_template('admin/group_discussions.html', items=items)


@admin_bp.route('/group-discussions/new', methods=['POST'])
@admin_required
def group_discussion_new():
    """新增小組討論"""
    data = request.get_json(silent=True) or {}
    title        = (data.get('title') or '本週小組討論').strip()
    youtube_url  = (data.get('youtube_url') or '').strip()
    questions    = data.get('questions') or []
    display_start = data.get('display_start', '')
    display_end   = data.get('display_end', '')
    is_active     = bool(data.get('is_active', True))

    if not youtube_url or not display_start or not display_end:
        return jsonify({'error': '請填寫 YouTube 連結與顯示日期區間'}), 400

    row = supabase.table('group_discussions').insert({
        'title':         title,
        'youtube_url':   youtube_url,
        'questions':     questions,
        'display_start': display_start,
        'display_end':   display_end,
        'is_active':     is_active,
    }).execute()
    return jsonify({'success': True, 'id': row.data[0]['id'] if row.data else None})


@admin_bp.route('/group-discussions/<gd_id>/edit', methods=['POST'])
@admin_required
def group_discussion_edit(gd_id):
    """更新小組討論"""
    data = request.get_json(silent=True) or {}
    patch = {}
    if 'title'         in data: patch['title']         = (data['title'] or '本週小組討論').strip()
    if 'youtube_url'   in data: patch['youtube_url']   = (data['youtube_url'] or '').strip()
    if 'questions'     in data: patch['questions'] = data['questions']
    if 'display_start' in data: patch['display_start'] = data['display_start']
    if 'display_end'   in data: patch['display_end']   = data['display_end']
    if 'is_active'     in data: patch['is_active']     = bool(data['is_active'])

    supabase.table('group_discussions').update(patch).eq('id', gd_id).execute()
    return jsonify({'success': True})


@admin_bp.route('/group-discussions/<gd_id>/delete', methods=['POST'])
@admin_required
def group_discussion_delete(gd_id):
    """刪除小組討論"""
    supabase.table('group_discussions').delete().eq('id', gd_id).execute()
    return jsonify({'success': True})


# ── 首頁卡片開關（超管） ──

@admin_bp.route('/portal-card-settings/toggle', methods=['POST'])
@super_admin_required
def portal_card_settings_toggle():
    """超管切換固定卡片可見度"""
    if not session.get('is_super_admin'):
        return jsonify({'error': '無權限'}), 403
    data = request.get_json(silent=True) or {}
    key = data.get('key', '').strip()
    visible = bool(data.get('visible', True))
    if not key:
        return jsonify({'error': '缺少 key'}), 400
    # upsert
    supabase.table('portal_card_settings').upsert(
        {'key': key, 'is_visible': visible},
        on_conflict='key'
    ).execute()
    return jsonify({'success': True, 'key': key, 'is_visible': visible})


@admin_bp.route('/portal-links/<link_id>/toggle', methods=['POST'])
@super_admin_required
def portal_link_toggle(link_id):
    """超管切換快捷連結可見度"""
    if not session.get('is_super_admin'):
        return jsonify({'error': '無權限'}), 403
    data = request.get_json(silent=True) or {}
    active = bool(data.get('is_active', True))
    supabase.table('portal_links').update({'is_active': active}).eq('id', link_id).execute()
    return jsonify({'success': True, 'is_active': active})


# ── 門戶卡片管理（超管）─────────────────────────────────────────

_PORTAL_CARDS_DEFAULT = [
    {'key': 'events',       'name': '活動報名',  'emoji': '🎉', 'subtitle': '查看並報名教會活動',     'url': '/events',                       'visible_to': 'all',         'is_active': True, 'sort_order': 10},
    {'key': 'calendar',     'name': '行事曆',    'emoji': '📅', 'subtitle': '教會行事曆與個人行程',   'url': '/calendar',                     'visible_to': 'member',      'is_active': True, 'sort_order': 20},
    {'key': 'bulletin',     'name': '每週週報',  'emoji': '📰', 'subtitle': '最新週報與公告',         'url': '/bulletins',                    'visible_to': 'all',         'is_active': True, 'sort_order': 30},
    {'key': 'prayer',       'name': '代禱牆',    'emoji': '🙏', 'subtitle': '分享需求，互相代禱',     'url': '/prayer',                       'visible_to': 'all',         'is_active': True, 'sort_order': 40},
    {'key': 'gospel',       'name': '福音探索',  'emoji': '✝️', 'subtitle': '認識信仰的第一步',       'url': '/gospel',                       'visible_to': 'all',         'is_active': True, 'sort_order': 50},
    {'key': 'diary',        'name': '天父日記',  'emoji': '📖', 'subtitle': '記錄每日與神的對話',     'url': '/diary',                        'visible_to': 'member',      'is_active': True, 'sort_order': 60},
    {'key': 'my_history',   'name': '電子簽到',  'emoji': '🗂️', 'subtitle': '我的活動出席紀錄',       'url': '/my-history',                   'visible_to': 'member',      'is_active': True, 'sort_order': 70},
    {'key': 'courses',      'name': '門訓學程',  'emoji': '📚', 'subtitle': '報名及追蹤進度',         'url': '/courses',                      'visible_to': 'member',      'is_active': True, 'sort_order': 80},
    {'key': 'cell_report',  'name': '小組回報',  'emoji': '👥', 'subtitle': '填寫本週小組聚會回報',   'url': '/cell-report/portal',           'visible_to': 'cell_leader', 'is_active': True, 'sort_order': 90},
    {'key': 'pastor_report','name': '牧者查閱',  'emoji': '📊', 'subtitle': '查看各小組回報與統計',   'url': '/cell-report/pastor-dashboard', 'visible_to': 'pastor',      'is_active': True, 'sort_order': 100},
    {'key': 'staff_report', 'name': '同工查閱',  'emoji': '📋', 'subtitle': '各區小組回報總覽',       'url': '/cell-report/staff-dashboard',  'visible_to': 'staff',       'is_active': True, 'sort_order': 110},
    {'key': 'pastor_diary', 'name': '查閱日記',  'emoji': '🔍', 'subtitle': '已授權的會友日記',       'url': '/diary/pastor',                 'visible_to': 'pastor',      'is_active': True, 'sort_order': 120},
    {'key': 'files',        'name': '檔案管理',  'emoji': '📁', 'subtitle': '教會資料夾與檔案',       'url': '/files',                        'visible_to': 'admin',       'is_active': True, 'sort_order': 130},
    {'key': 'admin',        'name': '後台管理',  'emoji': '⚙️', 'subtitle': '使用者、活動、系統設定', 'url': '/admin',                        'visible_to': 'admin',       'is_active': True, 'sort_order': 140},
]

PORTAL_CARDS_VISIBLE_TO_OPTIONS = [
    ('all',         '所有人（含訪客）'),
    ('member',      '已登入會員'),
    ('cell_leader', '小組長'),
    ('staff',       '同工'),
    ('pastor',      '牧者'),
    ('admin',       '管理員'),
]

PORTAL_CARDS_SETUP_SQL = """-- 在 Supabase SQL Editor 執行以下 SQL 建立門戶卡片資料表：
CREATE TABLE IF NOT EXISTS portal_cards (
  id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
  key TEXT UNIQUE NOT NULL,
  name TEXT NOT NULL,
  emoji TEXT DEFAULT '🔗',
  subtitle TEXT DEFAULT '',
  url TEXT NOT NULL,
  visible_to TEXT DEFAULT 'all',
  is_active BOOLEAN DEFAULT TRUE,
  sort_order INTEGER DEFAULT 0,
  is_system BOOLEAN DEFAULT TRUE,
  created_at TIMESTAMPTZ DEFAULT NOW()
);

INSERT INTO portal_cards (key, name, emoji, subtitle, url, visible_to, sort_order) VALUES
  ('events', '活動報名', '🎉', '查看並報名教會活動', '/events', 'all', 10),
  ('calendar', '行事曆', '📅', '教會行事曆與個人行程', '/calendar', 'member', 20),
  ('bulletin', '每週週報', '📰', '最新週報與公告', '/bulletins', 'all', 30),
  ('prayer', '代禱牆', '🙏', '分享需求，互相代禱', '/prayer', 'all', 40),
  ('gospel', '福音探索', '✝️', '認識信仰的第一步', '/gospel', 'all', 50),
  ('diary', '天父日記', '📖', '記錄每日與神的對話', '/diary', 'member', 60),
  ('my_history', '電子簽到', '🗂️', '我的活動出席紀錄', '/my-history', 'member', 70),
  ('courses', '門訓學程', '📚', '報名及追蹤進度', '/courses', 'member', 80),
  ('cell_report', '小組回報', '👥', '填寫本週小組聚會回報', '/cell-report/portal', 'cell_leader', 90),
  ('pastor_report', '牧者查閱', '📊', '查看各小組回報與統計', '/cell-report/pastor-dashboard', 'pastor', 100),
  ('staff_report', '同工查閱', '📋', '各區小組回報總覽', '/cell-report/staff-dashboard', 'staff', 110),
  ('pastor_diary', '查閱日記', '🔍', '已授權的會友日記', '/diary/pastor', 'pastor', 120),
  ('files', '檔案管理', '📁', '教會資料夾與檔案', '/files', 'admin', 130),
  ('admin', '後台管理', '⚙️', '使用者、活動、系統設定', '/admin', 'admin', 140)
ON CONFLICT (key) DO NOTHING;"""


def _load_portal_cards_from_db():
    """從 DB 載入門戶卡片，失敗時回傳預設值。
    回傳 (cards_list, from_db)"""
    try:
        rows = supabase.table('portal_cards').select('*').order('sort_order').execute().data or []
        if rows:
            return rows, True
    except Exception:
        logging.getLogger(__name__).warning('忽略非關鍵錯誤', exc_info=True)
    return _PORTAL_CARDS_DEFAULT, False


@admin_bp.route('/portal-cards')
@super_admin_required
def portal_cards_page():
    """門戶卡片管理頁面（僅超管）"""
    if not session.get('is_super_admin'):
        return redirect(url_for('admin.index'))
    cards, from_db = _load_portal_cards_from_db()
    links = supabase.table('portal_links').select('*').order('sort_order').execute().data or []
    return render_template('admin/portal_cards.html',
                           cards=cards,
                           from_db=from_db,
                           setup_sql=PORTAL_CARDS_SETUP_SQL,
                           visible_to_options=PORTAL_CARDS_VISIBLE_TO_OPTIONS,
                           links=links)


@admin_bp.route('/api/portal-cards/<key>', methods=['POST'])
@super_admin_required
def update_portal_card(key):
    """更新單張門戶卡片設定"""
    if not session.get('is_super_admin'):
        return jsonify({'error': '無權限'}), 403
    data = request.get_json(silent=True) or {}
    allowed = {'name', 'emoji', 'subtitle', 'visible_to', 'is_active', 'sort_order'}
    payload = {k: v for k, v in data.items() if k in allowed}
    if not payload:
        return jsonify({'error': '無有效欄位'}), 400
    try:
        supabase.table('portal_cards').update(payload).eq('key', key).execute()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/api/portal-cards/reorder', methods=['POST'])
@super_admin_required
def reorder_portal_cards():
    """批次更新 sort_order"""
    if not session.get('is_super_admin'):
        return jsonify({'error': '無權限'}), 403
    data = request.get_json(silent=True) or {}
    order_list = data.get('order', [])  # [{'key': 'events', 'sort_order': 10}, ...]
    try:
        for item in order_list:
            supabase.table('portal_cards')\
                .update({'sort_order': item['sort_order']})\
                .eq('key', item['key']).execute()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/portal-cards/group-defaults', methods=['GET', 'POST'])
@super_admin_required
def portal_group_defaults():
    """小組首頁預設：區塊顯示 + 卡片顯示"""
    import settings_store as ss, json as _json

    SECTIONS = [
        {'key': 'hero',                  'label': 'Hero 橫幅',        'emoji': '🖼️'},
        {'key': 'cell_report_reminder',  'label': '小組回報提醒',     'emoji': '📋'},
        {'key': 'attendance_summary',    'label': '聚會人數總覽',     'emoji': '📊'},
        {'key': 'staff_review',          'label': '同工 / 牧者查閱',  'emoji': '🔍'},
        {'key': 'upcoming_events',       'label': '我的近期活動',     'emoji': '📅'},
        {'key': 'diary_widget',          'label': '靈修空間',         'emoji': '📖'},
        {'key': 'portal_cards',          'label': '更多功能（功能磚）','emoji': '🔲'},
        {'key': 'group_discussion',      'label': '小組討論',         'emoji': '💬'},
        {'key': 'weekly_info',           'label': '本週資訊',         'emoji': '📰'},
    ]

    try:
        users_data = supabase.table('users').select('group_tags').execute().data or []
        all_tags = set()
        for u in users_data:
            for t in (u.get('group_tags') or []):
                if t:
                    all_tags.add(t)
        group_tags = sorted(all_tags)
    except Exception:
        group_tags = []

    try:
        portal_cards = supabase.table('portal_cards').select('key,name,emoji,subtitle')\
            .eq('is_active', True).order('sort_order').execute().data or []
    except Exception:
        portal_cards = []

    if request.method == 'POST':
        tag = request.form.get('group_tag', '').strip()
        if tag:
            hidden_sections = request.form.getlist('hidden_sections')
            ss.set(f'portal_sections_group_{tag}', _json.dumps({'hidden': hidden_sections}))
            hidden_cards = request.form.getlist('hidden_cards')
            ss.set(f'portal_group_{tag}', _json.dumps({'hidden': hidden_cards}))
        return redirect(url_for('admin.portal_group_defaults'))

    group_section_configs = {}
    group_card_configs = {}
    for tag in group_tags:
        raw_sec = ss.get(f'portal_sections_group_{tag}')
        group_section_configs[tag] = _json.loads(raw_sec) if raw_sec else {'hidden': []}
        raw_card = ss.get(f'portal_group_{tag}')
        group_card_configs[tag] = _json.loads(raw_card) if raw_card else {'hidden': []}

    return render_template('admin/portal_group_defaults.html',
        group_tags=group_tags, sections=SECTIONS, portal_cards=portal_cards,
        group_section_configs=group_section_configs,
        group_card_configs=group_card_configs)


def _has_finance_access():
    """超管永遠可以；或 group_tags 包含財務管理標籤之一；未設定標籤時退化為一般管理員"""
    if session.get('is_super_admin'):
        return True
    import settings_store as _ss
    raw = _ss.get('admin_finance_group_tags') or ''
    allowed = [t.strip() for t in raw.split(',') if t.strip()]
    if not allowed:
        return bool(session.get('is_admin'))
    user_tags = session.get('group_tags') or []
    return bool(set(user_tags) & set(allowed))


@admin_bp.route('/settings/payment', methods=['GET', 'POST'])
def payment_settings():
    if not _has_finance_access():
        return render_template('admin/forbidden.html'), 403
    import settings_store as ss
    saved = False
    if request.method == 'POST':
        ss.set('payment_gateway', request.form.get('gateway', 'none'))
        ss.set('payment_ecpay_merchant_id', request.form.get('ecpay_merchant_id', ''))
        ss.set('payment_ecpay_hash_key', request.form.get('ecpay_hash_key', ''))
        ss.set('payment_ecpay_hash_iv', request.form.get('ecpay_hash_iv', ''))
        ss.set('payment_ecpay_mode', request.form.get('ecpay_mode', 'test'))
        ss.set('payment_linepay_channel_id', request.form.get('linepay_channel_id', ''))
        ss.set('payment_linepay_channel_secret', request.form.get('linepay_channel_secret', ''))
        ss.set('payment_linepay_mode', request.form.get('linepay_mode', 'sandbox'))
        ss.set('payment_manual_instructions', request.form.get('manual_instructions', ''))
        ss.set('payment_fee_handling', request.form.get('fee_handling', 'church'))
        ss.set('payment_surcharge_rate', request.form.get('surcharge_rate', '3'))
        ss.set('payment_ecpay_credit_rate', request.form.get('ecpay_credit_rate', '2.75'))
        ss.set('payment_ecpay_credit_flat', request.form.get('ecpay_credit_flat', '1'))
        ss.set('payment_linepay_rate', request.form.get('linepay_rate', '2.9'))
        ss.set('payment_disclaimer', request.form.get('disclaimer', ''))
        if session.get('is_super_admin'):
            ss.set('admin_finance_group_tags', request.form.get('finance_group_tags', ''))
        saved = True
    settings = {
        'gateway': ss.get('payment_gateway') or 'none',
        'ecpay_merchant_id': ss.get('payment_ecpay_merchant_id') or '',
        'ecpay_hash_key': ss.get('payment_ecpay_hash_key') or '',
        'ecpay_hash_iv': ss.get('payment_ecpay_hash_iv') or '',
        'ecpay_mode': ss.get('payment_ecpay_mode') or 'test',
        'linepay_channel_id': ss.get('payment_linepay_channel_id') or '',
        'linepay_channel_secret': ss.get('payment_linepay_channel_secret') or '',
        'linepay_mode': ss.get('payment_linepay_mode') or 'sandbox',
        'manual_instructions': ss.get('payment_manual_instructions') or '',
        'fee_handling': ss.get('payment_fee_handling') or 'church',
        'surcharge_rate': ss.get('payment_surcharge_rate') or '3',
        'ecpay_credit_rate': ss.get('payment_ecpay_credit_rate') or '2.75',
        'ecpay_credit_flat': ss.get('payment_ecpay_credit_flat') or '1',
        'linepay_rate': ss.get('payment_linepay_rate') or '2.9',
        'disclaimer': ss.get('payment_disclaimer') or '',
        'finance_group_tags': ss.get('admin_finance_group_tags') or '',
    }
    return render_template('admin/payment_settings.html', settings=settings, saved=saved)


# ══════════════════════════════════════════
# 收款對帳報表
# ══════════════════════════════════════════

@admin_bp.route('/payments')
def payment_ledger():
    if not _has_finance_access():
        return render_template('admin/forbidden.html'), 403

    import settings_store as ss
    from datetime import datetime, timezone, timedelta

    # 篩選條件
    filter_event = request.args.get('event_id', '')
    filter_status = request.args.get('status', 'all')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    try:
        # 撈所有有費用的活動（供篩選下拉）
        events_result = supabase.table('events').select('id, title, fee')\
            .gt('fee', 0).order('created_at', desc=True).execute()
        all_events = events_result.data or []

        # 基本查詢：已報名（含付款/未付款）
        query = supabase.table('registrations')\
            .select('id, event_id, user_id, payment_status, created_at')\
            .neq('status', 'cancelled')
        if filter_event:
            query = query.eq('event_id', filter_event)
        if filter_status != 'all':
            query = query.eq('payment_status', filter_status)
        if date_from:
            query = query.gte('created_at', date_from + 'T00:00:00+00:00')
        if date_to:
            query = query.lte('created_at', date_to + 'T23:59:59+00:00')

        regs_raw = query.order('created_at', desc=True).limit(500).execute().data or []

        # 批次撈活動與使用者
        event_ids = list({r['event_id'] for r in regs_raw if r.get('event_id')})
        event_map = {}
        if event_ids:
            ev_rows = supabase.table('events').select('id, title, fee').in_('id', event_ids).execute().data or []
            event_map = {e['id']: e for e in ev_rows}

        uid_set = list({r['user_id'] for r in regs_raw if r.get('user_id')})
        user_map = {}
        if uid_set:
            u_rows = supabase.table('users').select('id, real_name, display_name').in_('id', uid_set).execute().data or []
            user_map = {u['id']: u.get('real_name') or u.get('display_name') or '—' for u in u_rows}

        regs = []
        for r in regs_raw:
            ev = event_map.get(r.get('event_id') or '')
            if ev and (ev.get('fee') or 0) > 0:
                r['_event'] = ev
                r['_user_name'] = user_map.get(r.get('user_id') or '', '—')
                regs.append(r)

        gateway = ss.get('payment_gateway') or 'none'
        fee_handling = ss.get('payment_fee_handling') or 'church'
        ecpay_rate = float(ss.get('payment_ecpay_credit_rate') or 2.75)
        ecpay_flat = float(ss.get('payment_ecpay_credit_flat') or 1)
        linepay_rate = float(ss.get('payment_linepay_rate') or 2.9)
        taipei_tz = timezone(timedelta(hours=8))

        def estimate_fee(amount, gw):
            if gw == 'ecpay':
                return round(amount * ecpay_rate / 100 + ecpay_flat, 1)
            elif gw == 'linepay':
                return round(amount * linepay_rate / 100, 1)
            return 0

        rows = []
        total_charged = 0
        total_fee_est = 0
        total_paid = 0

        for r in regs:
            ev = r.get('_event') or {}
            fee = int(ev.get('fee') or 0)
            if fee <= 0:
                continue
            note = r.get('payment_note') or ''
            pay_gw = 'ecpay' if note.startswith('ecpay:') else \
                     'linepay' if note.startswith('linepay:') else \
                     gateway if r.get('payment_status') == 'paid' else 'manual'
            trade_no = note.split(':', 1)[1] if ':' in note else '—'
            is_paid = r.get('payment_status') == 'paid'
            fee_est = estimate_fee(fee, pay_gw) if is_paid else 0
            net = round(fee - fee_est, 1) if is_paid else 0
            raw_time = r.get('created_at', '')
            try:
                dt = datetime.fromisoformat(raw_time.replace('Z', '+00:00')).astimezone(taipei_tz)
                paid_at = dt.strftime('%Y/%m/%d %H:%M')
            except Exception:
                paid_at = raw_time[:16]
            rows.append({
                'reg_id': r['id'],
                'event_id': r['event_id'],
                'event_title': ev.get('title', '—'),
                'name': r.get('_user_name', '—'),
                'fee': fee,
                'payment_status': r.get('payment_status', 'unpaid'),
                'gateway': pay_gw,
                'trade_no': trade_no,
                'fee_est': fee_est,
                'net': net,
                'created_at': paid_at,
            })
            if is_paid:
                total_paid += 1
                total_charged += fee
                total_fee_est += fee_est

        total_net = round(total_charged - total_fee_est, 1)

        event_summary = {}
        for row in rows:
            eid = row['event_id']
            if eid not in event_summary:
                event_summary[eid] = {'title': row['event_title'], 'fee': row['fee'],
                                      'paid': 0, 'unpaid': 0, 'total_charged': 0, 'total_fee': 0}
            s = event_summary[eid]
            if row['payment_status'] == 'paid':
                s['paid'] += 1
                s['total_charged'] += row['fee']
                s['total_fee'] += row['fee_est']
            else:
                s['unpaid'] += 1
        for s in event_summary.values():
            s['net'] = round(s['total_charged'] - s['total_fee'], 1)

        return render_template('admin/payment_ledger.html',
            rows=rows, all_events=all_events,
            filter_event=filter_event, filter_status=filter_status,
            date_from=date_from, date_to=date_to,
            gateway=gateway, fee_handling=fee_handling,
            total_charged=total_charged,
            total_fee_est=round(total_fee_est, 1),
            total_net=total_net, total_paid=total_paid,
            event_summary=event_summary,
        )

    except Exception as e:
        print(f'[payment_ledger] error: {e}')
        import traceback; traceback.print_exc()
        return render_template('admin/payment_ledger.html',
            rows=[], all_events=[], error=str(e),
            filter_event=filter_event, filter_status=filter_status,
            date_from=date_from, date_to=date_to,
            gateway='none', fee_handling='church',
            total_charged=0, total_fee_est=0, total_net=0,
            total_paid=0, event_summary={},
        )


@admin_bp.route('/payments/export')
def payment_ledger_export():
    if not _has_finance_access():
        return render_template('admin/forbidden.html'), 403

    import csv, io, settings_store as ss
    from datetime import datetime, timezone, timedelta

    filter_event = request.args.get('event_id', '')
    filter_status = request.args.get('status', 'all')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    query = supabase.table('registrations')\
        .select('id, event_id, user_id, payment_status, created_at')\
        .neq('status', 'cancelled')
    if filter_event:
        query = query.eq('event_id', filter_event)
    if filter_status != 'all':
        query = query.eq('payment_status', filter_status)
    if date_from:
        query = query.gte('created_at', date_from + 'T00:00:00+00:00')
    if date_to:
        query = query.lte('created_at', date_to + 'T23:59:59+00:00')

    regs_result = query.order('created_at', desc=True).limit(2000).execute()
    regs_raw = regs_result.data or []

    # 批次撈活動與使用者
    ev_ids = list({r['event_id'] for r in regs_raw if r.get('event_id')})
    ev_map = {}
    if ev_ids:
        ev_rows = supabase.table('events').select('id, title, fee').in_('id', ev_ids).execute().data or []
        ev_map = {e['id']: e for e in ev_rows}
    uid_set = list({r['user_id'] for r in regs_raw if r.get('user_id')})
    u_map = {}
    if uid_set:
        u_rows = supabase.table('users').select('id, real_name, display_name').in_('id', uid_set).execute().data or []
        u_map = {u['id']: u.get('real_name') or u.get('display_name') or '—' for u in u_rows}

    regs = []
    for r in regs_raw:
        ev = ev_map.get(r.get('event_id') or '')
        if ev and (ev.get('fee') or 0) > 0:
            r['_event'] = ev
            r['_user_name'] = u_map.get(r.get('user_id') or '', '—')
            regs.append(r)

    gateway = ss.get('payment_gateway') or 'none'
    ecpay_rate = float(ss.get('payment_ecpay_credit_rate') or 2.75)
    ecpay_flat = float(ss.get('payment_ecpay_credit_flat') or 1)
    linepay_rate = float(ss.get('payment_linepay_rate') or 2.9)
    taipei_tz = timezone(timedelta(hours=8))

    def estimate_fee(amount, gw):
        if gw == 'ecpay':
            return round(amount * ecpay_rate / 100 + ecpay_flat, 1)
        elif gw == 'linepay':
            return round(amount * linepay_rate / 100, 1)
        return 0

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['報名時間', '活動名稱', '姓名', '活動費用(NT$)', '繳費狀態', '金流平台', '交易編號', '預估手續費(NT$)', '預估實收(NT$)'])

    for r in regs:
        ev = r.get('_event') or {}
        fee = int(ev.get('fee') or 0)
        note = r.get('payment_note') or ''
        pay_gw = 'ecpay' if note.startswith('ecpay:') else \
                 'linepay' if note.startswith('linepay:') else gateway
        trade_no = note.split(':', 1)[1] if ':' in note else ''
        is_paid = r.get('payment_status') == 'paid'
        fee_est = estimate_fee(fee, pay_gw) if is_paid else ''
        net = round(fee - fee_est, 1) if is_paid else ''
        raw_time = r.get('created_at', '')
        try:
            dt = datetime.fromisoformat(raw_time.replace('Z', '+00:00')).astimezone(taipei_tz)
            paid_at = dt.strftime('%Y/%m/%d %H:%M')
        except Exception:
            paid_at = raw_time[:16]
        status_label = {'paid': '已付款', 'unpaid': '未付款', 'waived': '免收費'}.get(r.get('payment_status', 'unpaid'), '未付款')
        gw_label = {'ecpay': '綠界科技', 'linepay': 'LINE Pay', 'manual': '手動'}.get(pay_gw, '—')
        writer.writerow([paid_at, ev.get('title', ''), r.get('_user_name', '—'), fee, status_label, gw_label, trade_no, fee_est, net])

    output.seek(0)
    from flask import Response
    return Response(
        '﻿' + output.getvalue(),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': 'attachment; filename=payment_ledger.csv'}
    )


# ══════════════════════════════════════════
# 資料交換中心後台
# ══════════════════════════════════════════

@admin_bp.route('/files')
@admin_required
def files_admin():
    """資料交換中心管理後台"""
    import settings_store

    # 總用量
    all_files = supabase.table('files').select('id, name, file_size, owner_id, created_at, visibility').execute().data or []
    total_used = sum(f.get('file_size') or 0 for f in all_files)
    max_bytes = settings_store.get_max_storage_bytes()
    max_gb = int(settings_store.get('max_storage_gb') or 10)

    # 每位用戶用量
    user_usage = {}
    for f in all_files:
        oid = f.get('owner_id') or 'unknown'
        user_usage[oid] = user_usage.get(oid, 0) + (f.get('file_size') or 0)

    # 取得用戶名稱
    user_ids = [uid for uid in user_usage if uid != 'unknown']
    user_map = {}
    if user_ids:
        urows = supabase.table('users').select('id, real_name, display_name').in_('id', user_ids).execute().data or []
        for u in urows:
            user_map[u['id']] = u.get('real_name') or u.get('display_name') or '未知'

    user_stats = sorted([
        {'id': uid, 'name': user_map.get(uid, uid[:8] + '…'), 'bytes': sz}
        for uid, sz in user_usage.items()
    ], key=lambda x: x['bytes'], reverse=True)

    # 最近上傳（20筆）
    recent_files = sorted(all_files, key=lambda f: f.get('created_at') or '', reverse=True)[:20]
    for f in recent_files:
        oid = f.get('owner_id', '')
        f['owner_name'] = user_map.get(oid, '未知')

    return render_template('admin/files_admin.html',
                           total_used=total_used,
                           max_bytes=max_bytes,
                           max_gb=max_gb,
                           user_stats=user_stats,
                           recent_files=recent_files,
                           file_count=len(all_files))


@admin_bp.route('/api/files/set-limit', methods=['POST'])
@admin_required
def files_set_limit():
    """設定全域儲存上限"""
    data = request.get_json(silent=True) or {}
    try:
        gb = int(data.get('max_gb') or 10)
        if gb < 1 or gb > 1000:
            return jsonify({'error': 'GB 需介於 1～1000'}), 400
    except (ValueError, TypeError):
        return jsonify({'error': '請輸入有效數字'}), 400
    import settings_store
    settings_store.set('max_storage_gb', gb)
    return jsonify({'success': True, 'max_gb': gb})


@admin_bp.route('/api/files/<file_id>/delete', methods=['POST'])
@admin_required
def files_admin_delete(file_id):
    """管理員強制刪除任意檔案"""
    from storage import delete_file as r2_delete
    row = supabase.table('files').select('file_key, name').eq('id', file_id).execute().data
    if not row:
        return jsonify({'error': '檔案不存在'}), 404
    file_key = row[0]['file_key']
    try:
        r2_delete(file_key)
    except Exception:
        logging.getLogger(__name__).warning('忽略非關鍵錯誤', exc_info=True)
    supabase.table('files').delete().eq('id', file_id).execute()
    return jsonify({'success': True})
