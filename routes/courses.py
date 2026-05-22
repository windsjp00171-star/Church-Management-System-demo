# 門訓學程模組路由
from flask import Blueprint, session, redirect, url_for, render_template, request, jsonify, Response
from db import supabase
import secrets
import io
import hmac
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timezone, timedelta
from routes.notifications import create_notification
from routes.decorators import login_required, admin_required, super_admin_required
courses_bp = Blueprint('courses', __name__)

TAIPEI_TZ = timezone(timedelta(hours=8))

def _taipei_to_utc(s):
    """datetime-local 輸入值（台灣時間）→ UTC ISO 字串存入 DB"""
    if not s:
        return None
    try:
        dt = datetime.fromisoformat(str(s))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=TAIPEI_TZ)
        return dt.astimezone(timezone.utc).isoformat()
    except Exception:
        return s


@courses_bp.before_request
def guard_courses_admin():
    """
    /admin/courses/* 路由：需要超級管理員
    /admin/certifications/* 路由：需要超級管理員
    其他課程路由（學員端）：僅需登入
    """
    if request.path.startswith('/admin/'):
        if not session.get('user_id'):
            session['next_url'] = request.path
            return redirect(url_for('auth.login_page'))
        if not session.get('is_admin'):
            return render_template('admin/forbidden.html'), 403

    # CSRF 保護（只攔後台 POST 等）
    if request.method in ('POST', 'PUT', 'DELETE', 'PATCH') \
            and request.path.startswith('/admin/'):
        token = request.headers.get('X-CSRF-Token') or request.form.get('_csrf_token')
        expected = session.get('_csrf_token')
        if not expected or not token or not hmac.compare_digest(token, expected):
            return jsonify({'error': 'CSRF token 驗證失敗，請重新整理頁面'}), 403


# ── 權限裝飾器 ──────────────────────────────────────────





# ── 工具函式 ──────────────────────────────────────────

def _material_auto_out(course_id, enrollment_id):
    """報名且需要教材時，依課程學程自動扣庫存一筆"""
    try:
        course = supabase.table('courses').select('category_id, name').eq('id', course_id).execute()
        if not course.data:
            return
        cat_id = course.data[0].get('category_id')
        if not cat_id:
            return
        mat = supabase.table('materials').select('id, name, unit').eq('category_id', cat_id).eq('is_active', True).limit(1).execute()
        if not mat.data:
            return
        m = mat.data[0]
        supabase.table('material_transactions').insert({
            'material_id': m['id'],
            'type': 'out',
            'quantity': 1,
            'unit_cost': 0,
            'total_cost': 0,
            'course_id': course_id,
            'enrollment_id': enrollment_id,
            'notes': '系統自動－報名',
            'transaction_date': datetime.now(TAIPEI_TZ).date().isoformat(),
        }).execute()
    except Exception:
        pass  # 庫存扣除失敗不阻斷報名


def _material_auto_return(enrollment_id):
    """取消報名時，找回對應的自動扣庫存並退回一筆"""
    try:
        txn = supabase.table('material_transactions')\
            .select('id, material_id, course_id')\
            .eq('enrollment_id', enrollment_id).eq('type', 'out')\
            .eq('notes', '系統自動－報名').limit(1).execute()
        if not txn.data:
            return
        t = txn.data[0]
        supabase.table('material_transactions').insert({
            'material_id': t['material_id'],
            'type': 'in',
            'quantity': 1,
            'unit_cost': 0,
            'total_cost': 0,
            'course_id': t['course_id'],
            'enrollment_id': enrollment_id,
            'notes': '系統自動－取消報名',
            'transaction_date': datetime.now(TAIPEI_TZ).date().isoformat(),
        }).execute()
    except Exception:
        pass


def search_users(q, extra_fields='', limit=20):
    """共用用戶搜尋：依真實姓名或 LINE 暱稱，合併去重"""
    fields = f'id, real_name, display_name, picture_url, group_tags{", " + extra_fields if extra_fields else ""}'
    if not q:
        return []
    by_real = supabase.table('users').select(fields).ilike('real_name', f'%{q}%').limit(limit).execute().data or []
    by_line = supabase.table('users').select(fields).ilike('display_name', f'%{q}%').limit(limit).execute().data or []
    seen, merged = set(), []
    for u in by_real + by_line:
        if u['id'] not in seen:
            seen.add(u['id'])
            merged.append(u)
    return merged[:limit]


def get_course_or_404(course_id):
    result = supabase.table('courses').select('*').eq('id', course_id).execute()
    return result.data[0] if result.data else None


def get_attendance_map(session_ids):
    """session_id → list of user_ids 已出席"""
    if not session_ids:
        return {}
    att = supabase.table('session_attendance')\
        .select('session_id, user_id')\
        .in_('session_id', session_ids).execute()
    result = {}
    for a in (att.data or []):
        result.setdefault(a['session_id'], [])
        if a['user_id'] not in result[a['session_id']]:
            result[a['session_id']].append(a['user_id'])
    return result


def check_and_auto_complete(course_id, user_id):
    """簽到後檢查是否達到完訓門檻，達標則自動標記完訓，並嘗試發學程認證。"""
    course = supabase.table('courses')\
        .select('absence_limit, category_id').eq('id', course_id).single().execute().data
    if not course:
        return False

    enrollment = supabase.table('course_enrollments')\
        .select('id, status')\
        .eq('course_id', course_id).eq('user_id', user_id).single().execute().data
    if not enrollment or enrollment['status'] == 'completed':
        return False

    sessions = supabase.table('course_sessions')\
        .select('id').eq('course_id', course_id).execute().data or []
    total = len(sessions)
    if total == 0:
        return False

    sids = [s['id'] for s in sessions]
    att = supabase.table('session_attendance')\
        .select('session_id')\
        .eq('user_id', user_id).in_('session_id', sids).execute().data or []
    attended = len(att)
    absent = total - attended

    absence_limit = course.get('absence_limit')
    if attended > 0 and (absence_limit is None or absent <= absence_limit):
        now = datetime.now(TAIPEI_TZ).isoformat()
        supabase.table('course_enrollments').update({
            'status': 'completed',
            'completed_at': now,
        }).eq('id', enrollment['id']).execute()

        # 課程完訓後，自動發學程認證（若尚未持有）
        _auto_grant_certification(course_id, user_id, course.get('category_id'), now)
        return True
    return False


def _auto_grant_certification(course_id, user_id, category_id, now):
    """課程完訓後，檢查並自動發該學程認證。同學程已有認證則略過。"""
    if not category_id:
        return

    # 已持有此學程認證 → 不重複寫
    existing = supabase.table('course_certificates')\
        .select('id').eq('user_id', user_id).eq('category_id', category_id).execute()
    if existing.data:
        return

    today = datetime.now(TAIPEI_TZ).strftime('%Y-%m-%d')
    supabase.table('course_certificates').insert({
        'user_id': user_id,
        'category_id': category_id,
        'certified_at': today,
        'note': '系統自動認證',
    }).execute()


# ══════════════════════════════════════════
# 後台：學程管理
# ══════════════════════════════════════════

def auto_close_expired_courses():
    """懶式執行：把 reg_deadline 已過期但仍 is_open=True 的學程自動關閉"""
    try:
        now_iso = datetime.now(TAIPEI_TZ).isoformat()
        supabase.table('courses')\
            .update({'is_open': False})\
            .eq('is_open', True)\
            .not_.is_('reg_deadline', 'null')\
            .lt('reg_deadline', now_iso)\
            .execute()
    except Exception:
        pass  # 失敗靜默，下次再試


@courses_bp.route('/admin/courses')
@admin_required
def admin_courses():
    """後台：學程列表（進入時順帶關閉已過期報名的學程）"""
    auto_close_expired_courses()
    result = supabase.table('courses').select('*').order('created_at', desc=True).execute()
    courses = result.data or []

    if courses:
        cids = [c['id'] for c in courses]
        enroll_result = supabase.table('course_enrollments')\
            .select('course_id')\
            .in_('course_id', cids)\
            .eq('status', 'enrolled')\
            .execute()
        count_map = {}
        for e in (enroll_result.data or []):
            count_map[e['course_id']] = count_map.get(e['course_id'], 0) + 1
        for c in courses:
            c['enrolled_count'] = count_map.get(c['id'], 0)

    return render_template('courses/admin_list.html', courses=courses)


@courses_bp.route('/admin/courses/new', methods=['GET', 'POST'])
@admin_required
def admin_course_new():
    """後台：建立學程"""
    if request.method == 'POST':
        data = request.get_json()
        use_shared = data.get('use_shared_checkin', False)
        course_data = {
            'title': data.get('title', '').strip(),
            'location': data.get('location', '').strip() or None,
            'description': (data.get('description') or '').strip() or None,
            'notes': (data.get('notes') or '').strip() or None,
            'meal_options': data.get('meal_options') or None,
            'period': data.get('period', '').strip() or None,
            'total_sessions': int(data['total_sessions']) if data.get('total_sessions') else 1,
            'absence_limit': int(data['absence_limit']) if data.get('absence_limit') not in (None, '', 'null') else None,
            'material_fee': int(data.get('material_fee') or 0),
            'reminder_days': int(data.get('reminder_days') or 3),
            'reg_deadline': data.get('reg_deadline') or None,
            'prerequisite_course_id': data.get('prerequisite_course_id') or None,
            'category_id': data.get('category_id') or None,
            'prerequisite_category_id': data.get('prerequisite_category_id') or None,
            'is_open': data.get('is_open', True),
            'has_material': data.get('has_material', False),
            'whitelist_enabled': data.get('whitelist_enabled', False),
            'auto_complete_on_checkin': data.get('auto_complete_on_checkin', False),
            'shared_checkin_token': secrets.token_urlsafe(16) if use_shared else None,
            'created_by': session.get('user_id'),
        }
        if not course_data['title']:
            return jsonify({'error': '請填寫學程名稱'}), 400
        result = supabase.table('courses').insert(course_data).execute()
        return jsonify({'success': True, 'course_id': result.data[0]['id']})

    # 撈所有學程供「前置學程」下拉選單 + 類別
    all_courses = supabase.table('courses').select('id, title').order('created_at', desc=True).execute()
    categories = supabase.table('course_categories').select('*').order('sort_order').execute().data or []
    return render_template('courses/admin_form.html', course=None, all_courses=all_courses.data or [], categories=categories)


@courses_bp.route('/admin/courses/<course_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_course_edit(course_id):
    """後台：編輯學程"""
    course = get_course_or_404(course_id)
    if not course:
        return '找不到此學程', 404

    if request.method == 'POST':
        data = request.get_json()
        use_shared = data.get('use_shared_checkin', False)
        # 若原本沒有 token 且現在要開啟，就產生一個新的
        existing_token = course.get('shared_checkin_token')
        if use_shared and not existing_token:
            existing_token = secrets.token_urlsafe(16)
        elif not use_shared:
            existing_token = None
        update_data = {
            'title': data.get('title', '').strip(),
            'location': data.get('location', '').strip() or None,
            'description': (data.get('description') or '').strip() or None,
            'notes': (data.get('notes') or '').strip() or None,
            'meal_options': data.get('meal_options') or None,
            'period': data.get('period', '').strip() or None,
            'total_sessions': int(data['total_sessions']) if data.get('total_sessions') else 1,
            'absence_limit': int(data['absence_limit']) if data.get('absence_limit') not in (None, '', 'null') else None,
            'material_fee': int(data.get('material_fee') or 0),
            'reminder_days': int(data.get('reminder_days') or 3),
            'reg_deadline': data.get('reg_deadline') or None,
            'prerequisite_course_id': data.get('prerequisite_course_id') or None,
            'category_id': data.get('category_id') or None,
            'prerequisite_category_id': data.get('prerequisite_category_id') or None,
            'is_open': data.get('is_open', True),
            'has_material': data.get('has_material', False),
            'whitelist_enabled': data.get('whitelist_enabled', False),
            'auto_complete_on_checkin': data.get('auto_complete_on_checkin', False),
            'shared_checkin_token': existing_token,
        }
        supabase.table('courses').update(update_data).eq('id', course_id).execute()
        return jsonify({'success': True})

    all_courses = supabase.table('courses').select('id, title')\
        .neq('id', course_id).order('created_at', desc=True).execute()
    categories = supabase.table('course_categories').select('*').order('sort_order').execute().data or []
    return render_template('courses/admin_form.html', course=course, all_courses=all_courses.data or [], categories=categories)


@courses_bp.route('/admin/courses/<course_id>/toggle', methods=['POST'])
@admin_required
def admin_course_toggle(course_id):
    """開關報名"""
    course = get_course_or_404(course_id)
    if not course:
        return jsonify({'error': '找不到此學程'}), 404
    new_val = not course['is_open']
    supabase.table('courses').update({'is_open': new_val}).eq('id', course_id).execute()
    return jsonify({'success': True, 'is_open': new_val})


@courses_bp.route('/admin/courses/<course_id>/delete', methods=['POST'])
@admin_required
def admin_course_delete(course_id):
    """刪除學程（含關聯資料）"""
    # 先撈所有堂次 id，用來刪出席紀錄
    sessions = supabase.table('course_sessions').select('id').eq('course_id', course_id).execute().data or []
    session_ids = [s['id'] for s in sessions]
    if session_ids:
        supabase.table('session_attendance').delete().in_('session_id', session_ids).execute()
    supabase.table('course_sessions').delete().eq('course_id', course_id).execute()
    supabase.table('course_enrollments').delete().eq('course_id', course_id).execute()
    supabase.table('courses').delete().eq('id', course_id).execute()
    return jsonify({'success': True})


# ── 堂次管理 ──────────────────────────────────────────

@courses_bp.route('/admin/courses/<course_id>')
@admin_required
def admin_course_detail(course_id):
    """後台：學程詳細（堂次列表）"""
    course = get_course_or_404(course_id)
    if not course:
        return '找不到此學程', 404

    sessions_result = supabase.table('course_sessions')\
        .select('*').eq('course_id', course_id).order('session_number').execute()
    sessions = sessions_result.data or []

    # 每堂出席人數
    if sessions:
        sids = [s['id'] for s in sessions]
        att_map = get_attendance_map(sids)
        for s in sessions:
            s['attended_count'] = len(att_map.get(s['id'], set()))

    # 報名人數（只計算目前 enrolled）
    enroll_result = supabase.table('course_enrollments')\
        .select('id').eq('course_id', course_id)\
        .eq('status', 'enrolled').execute()
    enrolled_count = len(enroll_result.data or [])

    return render_template('courses/admin_detail.html',
        course=course, sessions=sessions, enrolled_count=enrolled_count)


@courses_bp.route('/admin/courses/<course_id>/sessions', methods=['POST'])
@admin_required
def admin_session_add(course_id):
    """新增堂次"""
    data = request.get_json()
    session_number = int(data.get('session_number', 1))
    token = secrets.token_urlsafe(12)

    result = supabase.table('course_sessions').insert({
        'course_id': course_id,
        'session_number': session_number,
        'title': data.get('title', '').strip() or f'第 {session_number} 堂',
        'scheduled_at': _taipei_to_utc(data.get('scheduled_at')),
        'end_time': _taipei_to_utc(data.get('end_time')),
        'checkin_token': token,
    }).execute()
    return jsonify({'success': True, 'session': result.data[0]})


@courses_bp.route('/admin/courses/<course_id>/sessions/<session_id>/delete', methods=['POST'])
@admin_required
def admin_session_delete(course_id, session_id):
    """刪除堂次（含出席紀錄）"""
    supabase.table('session_attendance').delete().eq('session_id', session_id).execute()
    supabase.table('course_sessions').delete().eq('id', session_id).execute()
    return jsonify({'success': True})


@courses_bp.route('/admin/courses/<course_id>/sessions/<session_id>/edit', methods=['POST'])
@admin_required
def admin_session_edit(course_id, session_id):
    """編輯堂次資訊"""
    data = request.get_json() or {}
    update_data = {
        'session_number': int(data['session_number']) if data.get('session_number') else 1,
        'title': data.get('title', '').strip() or None,
        'scheduled_at': _taipei_to_utc(data.get('scheduled_at')),
        'end_time': _taipei_to_utc(data.get('end_time')),
    }
    supabase.table('course_sessions').update(update_data).eq('id', session_id).execute()
    return jsonify({'success': True})


@courses_bp.route('/admin/courses/<course_id>/sessions/<session_id>/qrcode')
@admin_required
def admin_session_qrcode(course_id, session_id):
    """堂次簽到 QR Code 頁"""
    course = get_course_or_404(course_id)
    sess_result = supabase.table('course_sessions').select('*').eq('id', session_id).execute()
    if not course or not sess_result.data:
        return '找不到此堂次', 404
    sess = sess_result.data[0]
    checkin_url = request.host_url.rstrip('/') + f"/course-checkin/{session_id}/{sess['checkin_token']}"
    return render_template('courses/session_qrcode.html', course=course, sess=sess, checkin_url=checkin_url)


# ── 出席名單 ──────────────────────────────────────────

@courses_bp.route('/admin/courses/<course_id>/roster')
@admin_required
def admin_course_roster(course_id):
    """後台：名單 + 出席狀況"""
    try:
        course = get_course_or_404(course_id)
        if not course:
            return '找不到此學程', 404

        # ── 本期報名名單（enrolled）──
        enrolled_raw = supabase.table('course_enrollments')\
            .select('*').eq('course_id', course_id)\
            .eq('status', 'enrolled')\
            .order('created_at', desc=True).execute().data or []

        # 去重（同一人只留最新一筆）
        seen_e = set()
        enrolled_records = []
        for e in enrolled_raw:
            if e['user_id'] not in seen_e:
                seen_e.add(e['user_id'])
                enrolled_records.append(e)

        enrolled_user_ids = {e['user_id'] for e in enrolled_records}

        # ── 歷史完訓名單（completed）──
        completed_raw = supabase.table('course_enrollments')\
            .select('*').eq('course_id', course_id)\
            .eq('status', 'completed')\
            .order('completed_at', desc=True).execute().data or []

        seen_c = set()
        completed_records = []
        for e in completed_raw:
            if e['user_id'] not in seen_c:
                seen_c.add(e['user_id'])
                completed_records.append(e)

        completed_user_ids = {e['user_id'] for e in completed_records}

        # ── 缺勤未通過名單（absent）──
        absent_raw = supabase.table('course_enrollments')\
            .select('*').eq('course_id', course_id)\
            .eq('status', 'absent')\
            .order('created_at', desc=True).execute().data or []

        seen_a = set()
        absent_records = []
        for e in absent_raw:
            if e['user_id'] not in seen_a:
                seen_a.add(e['user_id'])
                absent_records.append(e)

        # 回鍋 = 本期報名中、也有歷史完訓紀錄（同課程）或類別認證（course_certificates）的人
        returning_users = enrolled_user_ids & completed_user_ids

        # 補抓：透過 course_certificates 認定的回鍋（報名時走的就是這條路）
        if course.get('category_id') and enrolled_user_ids:
            cert_rows = supabase.table('course_certificates')\
                .select('user_id')\
                .eq('category_id', course['category_id'])\
                .in_('user_id', list(enrolled_user_ids)).execute().data or []
            certified_user_ids = {r['user_id'] for r in cert_rows}
            returning_users = returning_users | certified_user_ids

        # 需教材人數（本期）
        material_count = sum(1 for e in enrolled_records if e.get('needs_material'))

        # 堂次
        sessions = supabase.table('course_sessions')\
            .select('*').eq('course_id', course_id).order('session_number').execute().data or []

        # 所有使用者（供管理員選人用）
        all_users = supabase.table('users')\
            .select('id, real_name, display_name, picture_url, group_tags')\
            .order('real_name').execute().data or []

        # 使用者資料 map
        user_map = {u['id']: u for u in all_users}

        # 已報名的 user_id set（供新增學員頁標記）
        enrolled_ids = list(enrolled_user_ids)

        # 出席記錄 session_id → list(user_id)
        att_map = get_attendance_map([s['id'] for s in sessions])

        # 教材庫存狀態（供警告顯示）
        material_stock_info = None
        if course.get('has_material') and course.get('category_id'):
            mat = supabase.table('material_stock').select('name, stock, unit')\
                .eq('category_id', course['category_id']).eq('is_active', True).limit(1).execute()
            if mat.data:
                material_stock_info = mat.data[0]
                material_stock_info['demand'] = material_count

        return render_template('courses/admin_roster.html',
            course=course,
            enrolled_records=enrolled_records,
            completed_records=completed_records,
            absent_records=absent_records,
            sessions=sessions,
            user_map=user_map,
            all_users=all_users,
            enrolled_ids=enrolled_ids,
            att_map=att_map,
            returning_users=returning_users,
            completed_user_ids=completed_user_ids,
            material_count=material_count,
            material_stock_info=material_stock_info,
        )
    except Exception:
        return '名單頁載入失敗，請稍後再試', 500


@courses_bp.route('/admin/courses/<course_id>/roster/<enrollment_id>/mark-absent', methods=['POST'])
@admin_required
def admin_roster_mark_absent(course_id, enrollment_id):
    """標記學員缺勤未通過（enrolled → absent）"""
    supabase.table('course_enrollments').update({'status': 'absent'})\
        .eq('id', enrollment_id).eq('course_id', course_id).execute()
    return jsonify({'success': True})


@courses_bp.route('/admin/courses/<course_id>/roster/<enrollment_id>/restore', methods=['POST'])
@admin_required
def admin_roster_restore(course_id, enrollment_id):
    """還原缺勤/完訓記錄 → enrolled"""
    supabase.table('course_enrollments').update({'status': 'enrolled', 'completed_at': None})\
        .eq('id', enrollment_id).eq('course_id', course_id).execute()
    return jsonify({'success': True})


@courses_bp.route('/admin/courses/<course_id>/roster/<enrollment_id>/cancel', methods=['POST'])
@admin_required
def admin_cancel_enrollment(course_id, enrollment_id):
    """後台：取消學員報名（enrolled → dropped）"""
    result = supabase.table('course_enrollments')\
        .select('status, needs_material').eq('id', enrollment_id).eq('course_id', course_id).execute()
    if not result.data:
        return jsonify({'error': '找不到此報名紀錄'}), 404
    had_material = result.data[0].get('needs_material', False)
    supabase.table('course_enrollments').update({'status': 'dropped'})\
        .eq('id', enrollment_id).execute()
    if had_material:
        _material_auto_return(enrollment_id)
    return jsonify({'success': True})


@courses_bp.route('/admin/courses/<course_id>/roster/batch-complete', methods=['POST'])
@admin_required
def admin_batch_complete(course_id):
    """一鍵完課所有達標者（缺勤未超標 + 尚未完課）"""
    course = get_course_or_404(course_id)
    if not course:
        return jsonify({'error': '找不到此學程'}), 404

    # 撈所有 enrolled 狀態的報名
    enrollments = supabase.table('course_enrollments')\
        .select('id, user_id, status').eq('course_id', course_id)\
        .eq('status', 'enrolled').execute().data or []

    if not enrollments:
        return jsonify({'success': True, 'count': 0})

    # 撈所有堂次
    sessions = supabase.table('course_sessions')\
        .select('id').eq('course_id', course_id).execute().data or []
    total = len(sessions)
    absence_limit = course.get('absence_limit')

    # 撈出席記錄
    att_map = get_attendance_map([s['id'] for s in sessions])

    now = datetime.now(TAIPEI_TZ).isoformat()
    completed_ids = []
    for e in enrollments:
        attended = sum(1 for s in sessions if e['user_id'] in att_map.get(s['id'], set()))
        absent = total - attended
        # 達標條件：無缺勤限制，或缺勤數 <= 上限
        if absence_limit is None or absent <= absence_limit:
            completed_ids.append(e['id'])

    if completed_ids:
        supabase.table('course_enrollments').update({
            'status': 'completed',
            'completed_at': now,
        }).in_('id', completed_ids).execute()

    return jsonify({'success': True, 'count': len(completed_ids)})


@courses_bp.route('/admin/courses/<course_id>/roster/search-user')
@admin_required
def admin_search_user(course_id):
    """搜尋用戶供手動加入（依真實姓名或 LINE 暱稱）"""
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])
    return jsonify(search_users(q, limit=10))


@courses_bp.route('/admin/courses/<course_id>/roster/enroll', methods=['POST'])
@admin_required
def admin_manual_enroll(course_id):
    """手動將用戶加入學程（可同時標記完課）"""
    data = request.get_json() or {}
    user_id = data.get('user_id')
    mark_completed = data.get('mark_completed', False)
    if not user_id:
        return jsonify({'error': '請指定用戶'}), 400

    now = datetime.now(TAIPEI_TZ).isoformat()
    # 查課程教材費，決定預設付款狀態
    course_info = supabase.table('courses').select('material_fee').eq('id', course_id).execute()
    mat_fee = (course_info.data[0].get('material_fee') or 0) if course_info.data else 0
    default_payment = 'waived' if mat_fee == 0 else 'unpaid'

    # 只查目前 enrolled 狀態（避免多筆問題）
    existing = supabase.table('course_enrollments')\
        .select('id, status').eq('course_id', course_id).eq('user_id', user_id)\
        .eq('status', 'enrolled').execute()

    if existing.data:
        enroll_id = existing.data[0]['id']
        update = {'status': 'completed' if mark_completed else 'enrolled'}
        if mark_completed:
            update['completed_at'] = now
        supabase.table('course_enrollments').update(update).eq('id', enroll_id).execute()
    else:
        # 新增一筆（允許多筆，這是手動補錄或回鍋情境）
        insert = {
            'course_id': course_id,
            'user_id': user_id,
            'status': 'completed' if mark_completed else 'enrolled',
            'payment_status': default_payment,
        }
        if mark_completed:
            insert['completed_at'] = now
        supabase.table('course_enrollments').insert(insert).execute()

    return jsonify({'success': True})


@courses_bp.route('/admin/courses/<course_id>/roster/<enrollment_id>/update', methods=['POST'])
@admin_required
def admin_roster_update(course_id, enrollment_id):
    """後台：修改學員報名資料（教材、餐點選擇）"""
    data = request.get_json() or {}
    update = {}
    if 'needs_material' in data:
        update['needs_material'] = bool(data['needs_material'])
    if 'meal_selections' in data:
        meal_selections = data['meal_selections'] or []
        # 重新計算餐費
        meal_total = 0
        course = get_course_or_404(course_id)
        if course and meal_selections:
            meal_options_cfg = course.get('meal_options') or {}
            price_map = {o['id']: o.get('price', 0)
                         for o in meal_options_cfg.get('options', []) if o.get('enabled')}
            meal_total = sum(price_map.get(mid, 0) for mid in meal_selections)
        update['meal_selections'] = meal_selections or None
        update['meal_total'] = meal_total
    if not update:
        return jsonify({'error': '沒有要更新的資料'}), 400
    supabase.table('course_enrollments').update(update)\
        .eq('id', enrollment_id).eq('course_id', course_id).execute()
    return jsonify({'success': True})


@courses_bp.route('/admin/courses/<course_id>/roster/<enrollment_id>/payment', methods=['POST'])
@admin_required
def admin_roster_payment(course_id, enrollment_id):
    """切換課程學員繳費狀態：unpaid → paid → waived → unpaid"""
    result = supabase.table('course_enrollments')\
        .select('payment_status').eq('id', enrollment_id).eq('course_id', course_id).execute()
    if not result.data:
        return jsonify({'error': '找不到此報名記錄'}), 404
    current = result.data[0].get('payment_status', 'unpaid')
    cycle = {'unpaid': 'paid', 'paid': 'waived', 'waived': 'unpaid'}
    new_status = cycle.get(current, 'paid')
    supabase.table('course_enrollments').update({'payment_status': new_status})\
        .eq('id', enrollment_id).execute()
    return jsonify({'success': True, 'payment_status': new_status})


@courses_bp.route('/admin/courses/<course_id>/roster/<enrollment_id>/toggle-complete', methods=['POST'])
@admin_required
def admin_toggle_complete(course_id, enrollment_id):
    """手動標記完課 / 取消完課"""
    result = supabase.table('course_enrollments').select('status').eq('id', enrollment_id).execute()
    if not result.data:
        return jsonify({'error': '找不到此報名記錄'}), 404
    current = result.data[0]['status']
    if current == 'completed':
        new_status = 'enrolled'
        supabase.table('course_enrollments').update({
            'status': new_status, 'completed_at': None
        }).eq('id', enrollment_id).execute()
    else:
        new_status = 'completed'
        now = datetime.now(TAIPEI_TZ).isoformat()
        supabase.table('course_enrollments').update({
            'status': new_status, 'completed_at': now
        }).eq('id', enrollment_id).execute()
    return jsonify({'success': True, 'status': new_status})


@courses_bp.route('/admin/courses/<course_id>/sessions/<session_id>/attendance/<user_id>/toggle', methods=['POST'])
@admin_required
def admin_toggle_attendance(course_id, session_id, user_id):
    """手動切換出席狀態"""
    existing = supabase.table('session_attendance')\
        .select('id').eq('session_id', session_id).eq('user_id', user_id).execute()
    if existing.data:
        supabase.table('session_attendance').delete().eq('id', existing.data[0]['id']).execute()
        return jsonify({'success': True, 'attended': False})
    else:
        now = datetime.now(TAIPEI_TZ).isoformat()
        supabase.table('session_attendance').insert({
            'session_id': session_id,
            'user_id': user_id,
            'attended_at': now,
            'method': 'manual',
        }).execute()
        auto_completed = check_and_auto_complete(course_id, user_id)
        return jsonify({'success': True, 'attended': True, 'auto_completed': auto_completed})


# ── 小組門訓總覽 ──────────────────────────────────────────

@courses_bp.route('/admin/courses/group-overview')
@admin_required
def admin_group_overview():
    """後台：全教會小組門訓總覽（以完訓認證為依據）"""
    try:
        all_cats = supabase.table('course_categories').select('id, name, sort_order')\
            .eq('is_active', True).order('sort_order').execute().data or []
    except Exception:
        all_cats = []
    if not all_cats:
        return render_template('courses/group_overview.html',
            all_courses=[], group_names=[], group_data={}, no_group_users=[])

    # 從 cell_groups 取得所有活躍小組（作為門訓分組依據）
    cell_groups = supabase.table('cell_groups').select('id, name').eq('is_active', True).order('name').execute().data or []
    group_names = [g['name'] for g in cell_groups]
    cell_group_map = {g['id']: g['name'] for g in cell_groups}

    # 取得所有有連結 user_id 的活躍小組成員
    members_raw = supabase.table('cell_members').select('group_id, user_id, name')\
        .eq('is_active', True).not_.is_('user_id', 'null').execute().data or []

    # 所有完訓認證（從 course_certificates 撈）
    all_certs = supabase.table('course_certificates').select('user_id, category_id, certified_at').execute().data or []

    # 建立 cert_map: user_id → { category_id: certified_at }
    cert_map = {}
    for c in all_certs:
        cert_map.setdefault(c['user_id'], {})[c['category_id']] = c.get('certified_at', '')

    # 按小組分組
    group_data = {name: [] for name in group_names}
    placed_user_ids = set()

    for m in sorted(members_raw, key=lambda x: x.get('name') or ''):
        gname = cell_group_map.get(m['group_id'])
        if not gname:
            continue
        uid = m['user_id']
        if uid in placed_user_ids:
            continue  # 同一人在多個小組時只放一次
        placed_user_ids.add(uid)
        user_certs = {cat['id']: cert_map.get(uid, {}).get(cat['id']) for cat in all_cats}
        group_data[gname].append({
            'id': uid,
            'name': m.get('name') or '—',
            'courses': user_certs,
        })

    no_group_users = []  # 未分配到任何小組的用戶（保留欄位供模板使用）

    return render_template('courses/group_overview.html',
        all_courses=all_cats,   # 前端模板用 all_courses，這裡傳 categories
        group_names=group_names,
        group_data=group_data,
        no_group_users=no_group_users,
    )


@courses_bp.route('/admin/courses/group-overview/export')
@admin_required
def admin_group_overview_export():
    """匯出完訓總覽 Excel"""
    # 課程類別
    all_cats = supabase.table('course_categories').select('id, name, sort_order')\
        .eq('is_active', True).order('sort_order').execute().data or []

    # 從 cell_groups 取得所有活躍小組
    cell_groups = supabase.table('cell_groups').select('id, name').eq('is_active', True).order('name').execute().data or []
    group_names = [g['name'] for g in cell_groups]
    cell_group_map = {g['id']: g['name'] for g in cell_groups}

    # 取得所有有連結 user_id 的活躍小組成員
    members_raw = supabase.table('cell_members').select('group_id, user_id, name')\
        .eq('is_active', True).not_.is_('user_id', 'null').execute().data or []

    # 完訓認證
    all_certs = supabase.table('course_certificates').select('user_id, category_id, certified_at').execute().data or []
    cert_map = {}
    for c in all_certs:
        cert_map.setdefault(c['user_id'], {})[c['category_id']] = c.get('certified_at', '')

    # 按小組整理
    group_data = {name: [] for name in group_names}
    no_group_users = []
    placed_user_ids = set()
    for m in sorted(members_raw, key=lambda x: x.get('name') or ''):
        gname = cell_group_map.get(m['group_id'])
        if not gname:
            continue
        uid = m['user_id']
        if uid in placed_user_ids:
            continue
        placed_user_ids.add(uid)
        user_entry = {
            'name': m.get('name') or '—',
            'tags': [gname],
            'courses': {cat['id']: cert_map.get(uid, {}).get(cat['id']) for cat in all_cats},
        }
        group_data[gname].append(user_entry)

    # ── 建立 Excel ──
    wb = Workbook()
    ws = wb.active
    ws.title = '完訓總覽'

    # 樣式
    h_font  = Font(bold=True, color='FFFFFF')
    h_fill  = PatternFill('solid', fgColor='1565C0')
    g_fill  = PatternFill('solid', fgColor='E3F2FD')  # 小組標題列
    ok_fill = PatternFill('solid', fgColor='E8F5E9')
    center  = Alignment(horizontal='center', vertical='center')

    # 標題列
    headers = ['小組', '姓名'] + [cat['name'] for cat in all_cats]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = h_font
        cell.fill      = h_fill
        cell.alignment = center
    ws.row_dimensions[1].height = 22

    # 欄寬
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    for i in range(len(all_cats)):
        col_letter = ws.cell(row=1, column=3 + i).column_letter
        ws.column_dimensions[col_letter].width = max(10, len(all_cats[i]['name']) + 2)

    row_idx = 2
    # 各小組
    for g_name in group_names:
        members = group_data.get(g_name, [])
        if not members:
            continue
        # 小組標題列
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
        g_cell = ws.cell(row=row_idx, column=1, value=f'▶ {g_name}（{len(members)} 人）')
        g_cell.font = Font(bold=True, color='1565C0')
        g_cell.fill = g_fill
        row_idx += 1

        for m in members:
            ws.cell(row=row_idx, column=1, value=g_name)
            ws.cell(row=row_idx, column=2, value=m['name'])
            for col_i, cat in enumerate(all_cats, 3):
                val = m['courses'].get(cat['id'])
                cell = ws.cell(row=row_idx, column=col_i, value='✓' if val else '—')
                cell.alignment = center
                if val:
                    cell.fill = ok_fill
                    cell.font = Font(bold=True, color='2E7D32')
            row_idx += 1

    # 未分組
    if no_group_users:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
        ws.cell(row=row_idx, column=1, value=f'▶ 未分組（{len(no_group_users)} 人）').font = Font(bold=True, color='888888')
        row_idx += 1
        for m in no_group_users:
            ws.cell(row=row_idx, column=1, value='—')
            ws.cell(row=row_idx, column=2, value=m['name'])
            for col_i, cat in enumerate(all_cats, 3):
                val = m['courses'].get(cat['id'])
                cell = ws.cell(row=row_idx, column=col_i, value='✓' if val else '—')
                cell.alignment = center
                if val:
                    cell.fill = ok_fill
                    cell.font = Font(bold=True, color='2E7D32')
            row_idx += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = '完訓總覽.xlsx'
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename*=UTF-8\'\'{quote(filename)}'}
    )


# ══════════════════════════════════════════
# 後台：課程類別管理
# ══════════════════════════════════════════

@courses_bp.route('/admin/course-categories')
@admin_required
def admin_categories():
    """後台：課程類別列表"""
    cats = supabase.table('course_categories')\
        .select('*').order('sort_order').execute().data or []
    return render_template('courses/admin_categories.html', categories=cats)


@courses_bp.route('/admin/course-categories/new', methods=['POST'])
@admin_required
def admin_category_new():
    data = request.get_json() or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': '請填寫類別名稱'}), 400
    result = supabase.table('course_categories').insert({
        'name': name,
        'description': (data.get('description') or '').strip() or None,
        'sort_order': int(data.get('sort_order', 0)),
        'is_active': True,
    }).execute()
    return jsonify({'success': True, 'id': result.data[0]['id']})


@courses_bp.route('/admin/course-categories/<cat_id>/edit', methods=['POST'])
@admin_required
def admin_category_edit(cat_id):
    data = request.get_json() or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': '請填寫類別名稱'}), 400
    supabase.table('course_categories').update({
        'name': name,
        'description': (data.get('description') or '').strip() or None,
        'sort_order': int(data.get('sort_order', 0)),
        'is_active': data.get('is_active', True),
    }).eq('id', cat_id).execute()
    return jsonify({'success': True})


@courses_bp.route('/admin/course-categories/<cat_id>/delete', methods=['POST'])
@admin_required
def admin_category_delete(cat_id):
    # 檢查是否有學程使用此類別
    using = supabase.table('courses').select('id').eq('category_id', cat_id).limit(1).execute()
    if using.data:
        return jsonify({'error': '此類別已有學程使用，無法刪除'}), 400
    supabase.table('course_categories').delete().eq('id', cat_id).execute()
    return jsonify({'success': True})


# ══════════════════════════════════════════
# 後台：完訓認證管理
# ══════════════════════════════════════════

@courses_bp.route('/admin/certifications')
@admin_required
def admin_certifications():
    """後台：完訓認證管理 — 橫向對照表"""
    cats = supabase.table('course_categories')\
        .select('*').order('sort_order').execute().data or []
    active_cats = [c for c in cats if c.get('is_active', True)]

    q            = request.args.get('q', '').strip()
    group_filter = request.args.get('group', '').strip()
    tag_filter   = request.args.get('tag', '').strip()

    all_users_raw = []
    offset = 0
    batch  = 500
    while True:
        rows = supabase.table('users')\
            .select('id, real_name, display_name, picture_url, group_tags')\
            .range(offset, offset + batch - 1).execute().data or []
        all_users_raw.extend(rows)
        if len(rows) < batch:
            break
        offset += batch

    all_certs = supabase.table('course_certificates')\
        .select('*').execute().data or []

    cert_pivot = {}
    for c in all_certs:
        uid, cid = c['user_id'], c['category_id']
        if uid not in cert_pivot:
            cert_pivot[uid] = {}
        cert_pivot[uid][cid] = c

    for u in all_users_raw:
        u['_name']  = u.get('real_name') or u.get('display_name') or '—'
        tags        = u.get('group_tags') or []
        u['_tags']  = tags
        u['_group'] = tags[0] if tags else '未分組'

    all_tags   = sorted({tag for u in all_users_raw for tag in (u.get('group_tags') or [])})
    all_groups = sorted({u['_group'] for u in all_users_raw if u['_group'] != '未分組'})

    users = all_users_raw
    if q:
        users = [u for u in users if q.lower() in u['_name'].lower()]
    if tag_filter:
        users = [u for u in users if tag_filter in (u.get('group_tags') or [])]
    elif group_filter:
        users = [u for u in users if u['_group'] == group_filter]

    users = sorted(users, key=lambda x: (x['_group'], x['_name']))

    page      = max(1, int(request.args.get('page', 1)))
    page_size = 50
    total     = len(users)
    total_pages = max(1, (total + page_size - 1) // page_size)
    page      = min(page, total_pages)
    users_page = users[(page - 1) * page_size : page * page_size]

    return render_template('courses/admin_certifications.html',
        categories=active_cats, users=users_page, cert_pivot=cert_pivot,
        all_groups=all_groups, all_tags=all_tags,
        q=q, group_filter=group_filter, tag_filter=tag_filter,
        page=page, total_pages=total_pages, total=total, page_size=page_size)


@courses_bp.route('/admin/certifications/add', methods=['POST'])
@super_admin_required
def admin_certification_add():
    """新增完訓認證（可覆蓋已有的，更新日期與備註）"""
    data = request.get_json() or {}
    user_id = data.get('user_id')
    category_id = data.get('category_id')
    if not user_id or not category_id:
        return jsonify({'error': '請指定學員與類別'}), 400

    certified_at = data.get('certified_at') or datetime.now(TAIPEI_TZ).strftime('%Y-%m-%d')
    note = data.get('note', '').strip() or None

    # upsert：同一人同一類別只留一筆
    existing = supabase.table('course_certificates')\
        .select('id').eq('user_id', user_id).eq('category_id', category_id).execute()
    if existing.data:
        supabase.table('course_certificates').update({
            'certified_at': certified_at,
            'note': note,
            'created_by': session.get('user_id'),
        }).eq('id', existing.data[0]['id']).execute()
        return jsonify({'success': True, 'updated': True})
    else:
        supabase.table('course_certificates').insert({
            'user_id': user_id,
            'category_id': category_id,
            'certified_at': certified_at,
            'note': note,
            'created_by': session.get('user_id'),
        }).execute()
        return jsonify({'success': True, 'updated': False})


@courses_bp.route('/admin/certifications/<cert_id>/delete', methods=['POST'])
@super_admin_required
def admin_certification_delete(cert_id):
    supabase.table('course_certificates').delete().eq('id', cert_id).execute()
    return jsonify({'success': True})


@courses_bp.route('/admin/certifications/batch-candidates')
@admin_required
def admin_cert_batch_candidates():
    """回傳某類別下有完課記錄、但尚未認證的學員清單"""
    category_id = request.args.get('category_id', '').strip()
    if not category_id:
        return jsonify([])

    # 找出所屬此類別的所有課程
    courses_in_cat = supabase.table('courses').select('id')\
        .eq('category_id', category_id).execute().data or []
    course_ids = [c['id'] for c in courses_in_cat]
    if not course_ids:
        return jsonify([])

    # 找出這些課程中 status='completed' 的學員
    enrollments = supabase.table('course_enrollments')\
        .select('user_id')\
        .in_('course_id', course_ids)\
        .eq('status', 'completed').execute().data or []
    user_ids = list({e['user_id'] for e in enrollments if e.get('user_id')})
    if not user_ids:
        return jsonify([])

    # 已認證的排除掉
    already = supabase.table('course_certificates')\
        .select('user_id')\
        .eq('category_id', category_id)\
        .in_('user_id', user_ids).execute().data or []
    certified_ids = {r['user_id'] for r in already}
    pending_ids = [uid for uid in user_ids if uid not in certified_ids]
    if not pending_ids:
        return jsonify([])

    # 撈用戶資料
    users = supabase.table('users')\
        .select('id, real_name, display_name, picture_url')\
        .in_('id', pending_ids).execute().data or []

    return jsonify([{
        'id': u['id'],
        'name': u.get('real_name') or u.get('display_name') or '—',
        'picture_url': u.get('picture_url') or '',
    } for u in sorted(users, key=lambda x: x.get('real_name') or x.get('display_name') or '')])


@courses_bp.route('/admin/certifications/batch-add', methods=['POST'])
@super_admin_required
def admin_cert_batch_add():
    """批次新增完訓認證"""
    data = request.get_json() or {}
    category_id = data.get('category_id', '').strip()
    user_ids = data.get('user_ids', [])
    certified_at = data.get('certified_at') or datetime.now(TAIPEI_TZ).strftime('%Y-%m-%d')
    note = (data.get('note') or '').strip() or None

    if not category_id or not user_ids:
        return jsonify({'error': '請選擇類別與學員'}), 400

    success_count = 0
    for uid in user_ids:
        try:
            existing = supabase.table('course_certificates')\
                .select('id').eq('user_id', uid).eq('category_id', category_id).execute().data
            if existing:
                supabase.table('course_certificates').update({
                    'certified_at': certified_at, 'note': note,
                }).eq('id', existing[0]['id']).execute()
            else:
                supabase.table('course_certificates').insert({
                    'user_id': uid, 'category_id': category_id,
                    'certified_at': certified_at, 'note': note,
                }).execute()
            success_count += 1
        except Exception:
            pass

    return jsonify({'success': True, 'count': success_count})


@courses_bp.route('/admin/certifications/search-user')
@admin_required
def admin_cert_search_user():
    q = request.args.get('q', '').strip()
    if not q:
        # 空查詢：回傳所有用戶（依真實姓名排序）
        all_users = supabase.table('users')\
            .select('id, real_name, display_name, picture_url')\
            .order('real_name').limit(300).execute().data or []
        return jsonify(all_users)
    return jsonify(search_users(q, limit=30))


@courses_bp.route('/admin/certifications/bulk-save', methods=['POST'])
@super_admin_required
def admin_cert_bulk_save():
    """一次儲存多筆新增/刪除認證（勾選式 pivot table 用）"""
    data = request.get_json() or {}
    adds    = data.get('add', [])       # [{user_id, category_id, certified_at, note}, ...]
    deletes = data.get('delete', [])    # [cert_id, ...]

    add_count = del_count = 0

    for item in adds:
        uid  = (item.get('user_id') or '').strip()
        cid  = (item.get('category_id') or '').strip()
        date = item.get('certified_at') or datetime.now(TAIPEI_TZ).strftime('%Y-%m-%d')
        note = (item.get('note') or '').strip() or None
        if not uid or not cid:
            continue
        try:
            existing = supabase.table('course_certificates')\
                .select('id').eq('user_id', uid).eq('category_id', cid).execute().data
            if existing:
                supabase.table('course_certificates').update(
                    {'certified_at': date, 'note': note}
                ).eq('id', existing[0]['id']).execute()
            else:
                supabase.table('course_certificates').insert(
                    {'user_id': uid, 'category_id': cid, 'certified_at': date, 'note': note}
                ).execute()
            add_count += 1
        except Exception:
            pass

    for cert_id in deletes:
        try:
            supabase.table('course_certificates').delete().eq('id', cert_id).execute()
            del_count += 1
        except Exception:
            pass

    return jsonify({'success': True, 'added': add_count, 'deleted': del_count})


# ══════════════════════════════════════════
# 學員端：瀏覽與報名
# ══════════════════════════════════════════

@courses_bp.route('/courses')
@login_required
def course_list():
    """學員：學程列表"""
    result = supabase.table('courses').select('*').eq('is_open', True)\
        .order('created_at', desc=True).execute()
    courses = result.data or []

    # 撈前置學程名稱
    prereq_ids = list({c['prerequisite_course_id'] for c in courses if c.get('prerequisite_course_id')})
    prereq_map = {}
    if prereq_ids:
        prereqs = supabase.table('courses').select('id, title').in_('id', prereq_ids).execute()
        for p in (prereqs.data or []):
            prereq_map[p['id']] = p['title']

    # 我的報名狀態
    uid = session['user_id']
    my_enrollments = supabase.table('course_enrollments')\
        .select('course_id, status').eq('user_id', uid).execute().data or []
    my_map = {e['course_id']: e['status'] for e in my_enrollments}

    # 報名人數
    if courses:
        cids = [c['id'] for c in courses]
        enroll_result = supabase.table('course_enrollments')\
            .select('course_id').in_('course_id', cids)\
            .eq('status', 'enrolled').execute()
        count_map = {}
        for e in (enroll_result.data or []):
            count_map[e['course_id']] = count_map.get(e['course_id'], 0) + 1
        for c in courses:
            c['enrolled_count'] = count_map.get(c['id'], 0)
            c['my_status'] = my_map.get(c['id'])
            c['prereq_title'] = prereq_map.get(c.get('prerequisite_course_id'), '')

    return render_template('courses/list.html', courses=courses)


@courses_bp.route('/courses/<course_id>')
@login_required
def course_detail(course_id):
    """學員：學程詳細 + 報名"""
    course = get_course_or_404(course_id)
    if not course:
        return '找不到此學程', 404

    uid = session['user_id']

    # 我的報名狀態（優先 enrolled，其次最新的 completed；dropped 視為未報名）
    my_enroll = supabase.table('course_enrollments')\
        .select('*').eq('course_id', course_id).eq('user_id', uid)\
        .order('created_at', desc=True).execute()
    enrollment = None
    if my_enroll.data:
        for e in my_enroll.data:
            if e['status'] == 'enrolled':
                enrollment = e
                break
        if enrollment is None:
            latest = my_enroll.data[0]
            if latest['status'] != 'dropped':  # dropped 不顯示，讓頁面回到報名按鈕
                enrollment = latest

    # 堂次列表
    sessions = supabase.table('course_sessions')\
        .select('*').eq('course_id', course_id).order('session_number').execute().data or []

    # 我的出席記錄
    my_attendance = set()
    if sessions and enrollment:
        sids = [s['id'] for s in sessions]
        att = supabase.table('session_attendance')\
            .select('session_id').eq('user_id', uid).in_('session_id', sids).execute()
        my_attendance = {a['session_id'] for a in (att.data or [])}

    # 前置學程資格確認
    prereq_ok = True
    prereq_title = ''
    if course.get('prerequisite_course_id'):
        prereq = supabase.table('courses').select('title')\
            .eq('id', course['prerequisite_course_id']).execute()
        prereq_title = prereq.data[0]['title'] if prereq.data else ''
        # 檢查是否完課
        prereq_enroll = supabase.table('course_enrollments')\
            .select('status').eq('course_id', course['prerequisite_course_id'])\
            .eq('user_id', uid).execute()
        prereq_ok = bool(prereq_enroll.data and prereq_enroll.data[0]['status'] == 'completed')

    # 截止日判斷
    deadline_passed = False
    if course.get('reg_deadline'):
        deadline = datetime.fromisoformat(course['reg_deadline'].replace('Z', '+00:00'))
        deadline_passed = datetime.now(TAIPEI_TZ) > deadline.astimezone(TAIPEI_TZ)

    # 類別認證判斷
    user_certified = False     # 此用戶是否已有此課程類別的認證
    category = None
    if course.get('category_id'):
        cat_result = supabase.table('course_categories')\
            .select('*').eq('id', course['category_id']).execute()
        category = cat_result.data[0] if cat_result.data else None
        cert = supabase.table('course_certificates')\
            .select('id, certified_at').eq('user_id', uid)\
            .eq('category_id', course['category_id']).execute()
        user_certified = bool(cert.data)

    # 前置類別認證判斷（優先使用 prerequisite_category_id，fallback 舊版）
    prereq_category_name = ''
    if course.get('prerequisite_category_id'):
        pc = supabase.table('course_categories')\
            .select('name').eq('id', course['prerequisite_category_id']).execute()
        prereq_category_name = pc.data[0]['name'] if pc.data else ''
        cert2 = supabase.table('course_certificates')\
            .select('id').eq('user_id', uid)\
            .eq('category_id', course['prerequisite_category_id']).execute()
        prereq_ok = bool(cert2.data)
        prereq_title = prereq_category_name

    return render_template('courses/detail.html',
        course=course,
        enrollment=enrollment,
        sessions=sessions,
        my_attendance=my_attendance,
        prereq_ok=prereq_ok,
        prereq_title=prereq_title,
        deadline_passed=deadline_passed,
        user_certified=user_certified,
        category=category,
    )


@courses_bp.route('/courses/<course_id>/enroll', methods=['POST'])
@login_required
def course_enroll(course_id):
    """學員：報名學程"""
    course = get_course_or_404(course_id)
    if not course:
        return jsonify({'error': '找不到此學程'}), 404
    if not course.get('is_open'):
        return jsonify({'error': '此學程目前未開放報名'}), 400

    # 報名截止檢查
    if course.get('reg_deadline'):
        deadline = datetime.fromisoformat(course['reg_deadline'].replace('Z', '+00:00'))
        if datetime.now(TAIPEI_TZ) > deadline.astimezone(TAIPEI_TZ):
            return jsonify({'error': '報名已截止'}), 400

    uid = session['user_id']

    # 白名單檢查
    if course.get('whitelist_enabled'):
        wl = supabase.table('registration_whitelist')\
            .select('id').eq('ref_type', 'course').eq('ref_id', course_id)\
            .eq('user_id', uid).execute()
        if not wl.data:
            return jsonify({'error': '你不在此學程的報名名單中，請聯絡管理員'}), 403

    # 前置學程檢查
    if course.get('prerequisite_course_id'):
        prereq_enroll = supabase.table('course_enrollments')\
            .select('status').eq('course_id', course['prerequisite_course_id'])\
            .eq('user_id', uid).execute()
        if not prereq_enroll.data or prereq_enroll.data[0]['status'] != 'completed':
            return jsonify({'error': '尚未完成前置學程，無法報名'}), 400

    data_body = request.get_json() or {}
    force_enroll = bool(data_body.get('force_enroll', False))  # 回鍋確認後帶入
    needs_material = bool(data_body.get('needs_material', False))
    # 非回鍋且課程有教材費 → 強制購買，不信任前端
    if course.get('has_material') and not force_enroll:
        needs_material = True

    # 餐點選擇計算
    meal_selections = data_body.get('meal_selections') or []  # list of option IDs
    meal_total = 0
    meal_options_cfg = course.get('meal_options') or {}
    if meal_selections and meal_options_cfg.get('enabled'):
        price_map = {o['id']: o.get('price', 0)
                     for o in meal_options_cfg.get('options', []) if o.get('enabled')}
        meal_total = sum(price_map.get(mid, 0) for mid in meal_selections)

    # 前置類別認證檢查（優先 prerequisite_category_id）
    if course.get('prerequisite_category_id'):
        cert = supabase.table('course_certificates')\
            .select('id').eq('user_id', uid)\
            .eq('category_id', course['prerequisite_category_id']).execute()
        if not cert.data:
            # 撈類別名稱
            pc = supabase.table('course_categories')\
                .select('name').eq('id', course['prerequisite_category_id']).execute()
            pname = pc.data[0]['name'] if pc.data else '前置課程'
            return jsonify({'error': f'您尚未完成「{pname}」，無法報名此學程'}), 400

    # 類別認證回鍋提示（有認證但還沒 force_enroll）
    if course.get('category_id') and not force_enroll:
        cert = supabase.table('course_certificates')\
            .select('id, certified_at').eq('user_id', uid)\
            .eq('category_id', course['category_id']).execute()
        if cert.data:
            cat = supabase.table('course_categories')\
                .select('name').eq('id', course['category_id']).execute()
            cname = cat.data[0]['name'] if cat.data else '此課程'
            return jsonify({
                'already_certified': True,
                'certified_at': cert.data[0].get('certified_at', ''),
                'message': f'您已完訓「{cname}」，確定要回鍋再次參加嗎？',
            }), 200

    # 已有 enrolled 狀態 → 不重複報名
    active = supabase.table('course_enrollments')\
        .select('id').eq('course_id', course_id).eq('user_id', uid)\
        .eq('status', 'enrolled').execute()
    if active.data:
        return jsonify({'error': '你已報名此學程'}), 400

    # 計算總費用，費用為 0 → 自動標記免收費
    total_fee = (course.get('material_fee') or 0) * (1 if needs_material else 0) + meal_total
    auto_payment = 'waived' if total_fee == 0 else 'unpaid'

    # 有 dropped 狀態 → 重新啟用
    dropped = supabase.table('course_enrollments')\
        .select('id').eq('course_id', course_id).eq('user_id', uid)\
        .eq('status', 'dropped').execute()
    if dropped.data:
        enroll_id = dropped.data[0]['id']
        supabase.table('course_enrollments').update({
            'status': 'enrolled',
            'needs_material': needs_material,
            'meal_selections': meal_selections or None,
            'meal_total': meal_total,
            'payment_status': auto_payment,
        }).eq('id', enroll_id).execute()
        if needs_material:
            _material_auto_out(course_id, enroll_id)
        return jsonify({'success': True, 'message': '已重新報名'})

    # 新增一筆報名（回鍋也 insert 新筆，舊的 completed 保留）
    is_returning = force_enroll  # 由前端確認回鍋
    new_enroll = supabase.table('course_enrollments').insert({
        'course_id': course_id,
        'user_id': uid,
        'status': 'enrolled',
        'needs_material': needs_material,
        'meal_selections': meal_selections or None,
        'meal_total': meal_total,
        'payment_status': auto_payment,
    }).execute()
    if needs_material and new_enroll.data:
        _material_auto_out(course_id, new_enroll.data[0]['id'])
    msg = '已回鍋報名！歡迎再次參加 🔄' if is_returning else '報名成功！'

    # ── 報名成功通知 ──────────────────────────
    body_parts = []
    if course.get('has_material') and not force_enroll:
        fee = course.get('material_fee') or 0
        if fee:
            body_parts.append(f'教材費：${fee} 元，請記得繳費')
    create_notification(
        user_id  = uid,
        title    = f'📖 {"回鍋報名" if is_returning else "報名成功"} — {course["title"]}',
        body     = '\n'.join(body_parts) or None,
        type     = 'enrollment',
        link     = f'/courses/{course_id}',
        ref_type = 'course',
        ref_id   = course_id,
    )

    return jsonify({'success': True, 'message': msg})


@courses_bp.route('/courses/<course_id>/unenroll', methods=['POST'])
@login_required
def course_unenroll(course_id):
    """學員：取消報名"""
    uid = session['user_id']
    existing = supabase.table('course_enrollments')\
        .select('id, needs_material').eq('course_id', course_id).eq('user_id', uid)\
        .eq('status', 'enrolled').execute()
    if not existing.data:
        return jsonify({'error': '你尚未報名此學程'}), 400

    enroll_id = existing.data[0]['id']
    had_material = existing.data[0].get('needs_material', False)
    supabase.table('course_enrollments').update({'status': 'dropped'})\
        .eq('id', enroll_id).execute()
    if had_material:
        _material_auto_return(enroll_id)
    return jsonify({'success': True})


@courses_bp.route('/my-courses')
@login_required
def my_courses():
    """學員：我的學程"""
    uid = session['user_id']
    all_enrollments = supabase.table('course_enrollments')\
        .select('*').eq('user_id', uid)\
        .in_('status', ['enrolled', 'completed', 'absent'])\
        .order('created_at', desc=True).execute().data or []

    # 每個學程只顯示最活躍那筆（enrolled > absent > completed 優先序）
    priority = {'enrolled': 3, 'absent': 2, 'completed': 1}
    seen = {}
    for e in all_enrollments:
        cid = e['course_id']
        if cid not in seen or priority.get(e['status'], 0) > priority.get(seen[cid]['status'], 0):
            seen[cid] = e
    enrollments = list(seen.values())

    if not enrollments:
        return render_template('courses/my_courses.html', items=[])

    # 撈學程資料
    cids = [e['course_id'] for e in enrollments]
    courses_data = supabase.table('courses').select('*').in_('id', cids).execute().data or []
    course_map = {c['id']: c for c in courses_data}

    # 每個學程的堂次 + 我的出席數
    sessions_result = supabase.table('course_sessions')\
        .select('id, course_id').in_('course_id', cids).execute().data or []
    sessions_by_course = {}
    for s in sessions_result:
        sessions_by_course.setdefault(s['course_id'], []).append(s['id'])

    all_session_ids = [s['id'] for s in sessions_result]
    my_att = set()
    if all_session_ids:
        att_result = supabase.table('session_attendance')\
            .select('session_id').eq('user_id', uid)\
            .in_('session_id', all_session_ids).execute()
        my_att = {a['session_id'] for a in (att_result.data or [])}

    items = []
    for e in enrollments:
        course = course_map.get(e['course_id'], {})
        sids = sessions_by_course.get(e['course_id'], [])
        attended = len([s for s in sids if s in my_att])
        items.append({
            'enrollment': e,
            'course': course,
            'total_sessions': len(sids),
            'attended': attended,
        })

    # 我的完訓認證徽章
    certs_raw = supabase.table('course_certificates')\
        .select('*, course_categories(name, description)')\
        .eq('user_id', uid)\
        .order('certified_at', desc=True).execute().data or []

    return render_template('courses/my_courses.html', items=items, certifications=certs_raw)


# ══════════════════════════════════════════
# 學員端：課程通用 QR 簽到
# ══════════════════════════════════════════

@courses_bp.route('/course-shared-checkin/<token>')
@login_required
def course_shared_checkin(token):
    """通用 QR Code 簽到：整個課程共用一個 token，自動對應最近堂次"""
    # 找課程
    result = supabase.table('courses').select('*').eq('shared_checkin_token', token).execute()
    if not result.data:
        return render_template('courses/checkin_result.html',
            success=False, message='QR Code 無效或已過期')
    course = result.data[0]
    uid = session['user_id']

    # 確認已報名
    enrollment_result = supabase.table('course_enrollments')\
        .select('*').eq('course_id', course['id']).eq('user_id', uid).execute()
    if not enrollment_result.data or enrollment_result.data[0]['status'] == 'dropped':
        return render_template('courses/checkin_result.html',
            success=False, message='你尚未報名此學程，無法簽到')
    enrollment = enrollment_result.data[0]

    # 找最近的堂次（最接近今天的，優先未來場次，其次最近過去）
    now_str = datetime.now(TAIPEI_TZ).isoformat()
    sessions_result = supabase.table('course_sessions')\
        .select('*').eq('course_id', course['id']).order('scheduled_at').execute()
    sessions_list = sessions_result.data or []

    if not sessions_list:
        return render_template('courses/checkin_result.html',
            success=False, message='此學程尚未設定堂次，請聯絡管理員')

    # 找最近一堂（優先今天或最近的過去）
    target_sess = sessions_list[0]
    for s in sessions_list:
        if s.get('scheduled_at') and s['scheduled_at'] <= now_str:
            target_sess = s  # 持續更新，最後會是最近的過去堂次
    # 若全部都是未來，取第一堂
    if all((s.get('scheduled_at') or '') > now_str for s in sessions_list):
        target_sess = sessions_list[0]

    # 已簽到檢查
    existing = supabase.table('session_attendance')\
        .select('id').eq('session_id', target_sess['id']).eq('user_id', uid).execute()
    if existing.data:
        return render_template('courses/checkin_result.html',
            success=True, already=True,
            course=course, sess=target_sess,
            display_name=session.get('display_name'))

    # 簽到
    now = datetime.now(TAIPEI_TZ).isoformat()
    supabase.table('session_attendance').insert({
        'session_id': target_sess['id'],
        'user_id': uid,
        'attended_at': now,
        'method': 'qr',
    }).execute()

    auto_completed = check_and_auto_complete(course['id'], uid)

    return render_template('courses/checkin_result.html',
        success=True, already=False,
        course=course, sess=target_sess,
        display_name=session.get('display_name'),
        auto_completed=auto_completed)


# ══════════════════════════════════════════
# 學員端：掃碼簽到
# ══════════════════════════════════════════

@courses_bp.route('/course-checkin/<session_id>/<token>')
@login_required
def course_checkin(session_id, token):
    """學員掃碼簽到（GET 顯示確認頁，POST 完成簽到）"""
    sess_result = supabase.table('course_sessions').select('*').eq('id', session_id).execute()
    if not sess_result.data or sess_result.data[0].get('checkin_token') != token:
        return render_template('courses/checkin_result.html',
            success=False, message='QR Code 無效或已過期')

    sess = sess_result.data[0]
    course = get_course_or_404(sess['course_id'])
    uid = session['user_id']

    # 確認已報名
    enrollment = supabase.table('course_enrollments')\
        .select('status').eq('course_id', sess['course_id'])\
        .eq('user_id', uid).execute()
    if not enrollment.data or enrollment.data[0]['status'] == 'dropped':
        return render_template('courses/checkin_result.html',
            success=False, message='你尚未報名此學程，無法簽到')

    # 已簽到
    existing_att = supabase.table('session_attendance')\
        .select('id').eq('session_id', session_id).eq('user_id', uid).execute()
    if existing_att.data:
        return render_template('courses/checkin_result.html',
            success=True, already=True,
            course=course, sess=sess,
            display_name=session.get('display_name'))

    # 簽到
    now = datetime.now(TAIPEI_TZ).isoformat()
    supabase.table('session_attendance').insert({
        'session_id': session_id,
        'user_id': uid,
        'attended_at': now,
        'method': 'qr',
    }).execute()

    auto_completed = check_and_auto_complete(sess['course_id'], uid)

    return render_template('courses/checkin_result.html',
        success=True, already=False,
        course=course, sess=sess,
        display_name=session.get('display_name'),
        auto_completed=auto_completed)


# ══════════════════════════════════════════════════════
#  教材庫存管理
# ══════════════════════════════════════════════════════

@courses_bp.route('/admin/materials')
@admin_required
def admin_materials():
    """教材庫存總覽"""
    try:
        rows = supabase.table('material_stock').select('*').order('name').execute().data or []
    except Exception:
        rows = []
    try:
        categories = supabase.table('course_categories').select('id, name')\
            .eq('is_active', True).order('sort_order').execute().data or []
    except Exception:
        categories = []
    cat_map = {c['id']: c['name'] for c in categories}
    return render_template('courses/admin_materials.html',
        materials=rows, categories=categories, cat_map=cat_map)


@courses_bp.route('/admin/materials/new', methods=['POST'])
@admin_required
def admin_material_new():
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': '請輸入教材名稱'}), 400
    row = {
        'name': name,
        'description': (data.get('description') or '').strip() or None,
        'unit': (data.get('unit') or '本').strip(),
        'selling_price': int(data.get('selling_price') or 0),
        'category_id': data.get('category_id') or None,
    }
    result = supabase.table('materials').insert(row).execute()
    if not result.data:
        return jsonify({'error': '新增失敗'}), 500
    material_id = result.data[0]['id']

    # 若填了初始庫存，自動建一筆進貨記錄
    init_stock = int(data.get('init_stock') or 0)
    if init_stock > 0:
        init_cost = int(data.get('init_cost') or 0)
        supabase.table('material_transactions').insert({
            'material_id': material_id,
            'type': 'in',
            'quantity': init_stock,
            'unit_cost': init_cost,
            'total_cost': init_cost * init_stock,
            'notes': '初始盤點庫存',
            'transaction_date': datetime.now(TAIPEI_TZ).date().isoformat(),
        }).execute()

    return jsonify({'success': True, 'id': material_id})


@courses_bp.route('/admin/materials/<material_id>/edit', methods=['POST'])
@admin_required
def admin_material_edit(material_id):
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': '請輸入教材名稱'}), 400
    supabase.table('materials').update({
        'name': name,
        'description': (data.get('description') or '').strip() or None,
        'unit': (data.get('unit') or '本').strip(),
        'selling_price': int(data.get('selling_price') or 0),
        'category_id': data.get('category_id') or None,
    }).eq('id', material_id).execute()
    return jsonify({'success': True})


@courses_bp.route('/admin/materials/<material_id>/delete', methods=['POST'])
@admin_required
def admin_material_delete(material_id):
    data = request.get_json() or {}
    force = data.get('force', False)
    txns = supabase.table('material_transactions').select('id').eq('material_id', material_id).limit(1).execute()
    if txns.data and not force:
        return jsonify({'error': '此教材已有異動記錄，無法刪除'}), 400
    if txns.data and force:
        supabase.table('material_transactions').delete().eq('material_id', material_id).execute()
    supabase.table('materials').delete().eq('id', material_id).execute()
    return jsonify({'success': True})


@courses_bp.route('/admin/materials/<material_id>')
@admin_required
def admin_material_detail(material_id):
    """教材詳細頁：異動記錄"""
    mat = supabase.table('material_stock').select('*').eq('id', material_id).execute()
    if not mat.data:
        return redirect('/admin/materials')
    material = mat.data[0]

    txns = supabase.table('material_transactions').select('*')\
        .eq('material_id', material_id)\
        .order('transaction_date', desc=True)\
        .order('created_at', desc=True).execute().data or []

    # 附上課程名稱
    course_ids = list({t['course_id'] for t in txns if t.get('course_id')})
    course_map = {}
    if course_ids:
        cs = supabase.table('courses').select('id, name').in_('id', course_ids).execute().data or []
        course_map = {c['id']: c['name'] for c in cs}

    # 課程清單供銷售下拉
    all_courses = supabase.table('courses').select('id, title')\
        .eq('is_open', True).order('title').execute().data or []

    return render_template('courses/admin_material_detail.html',
        material=material, transactions=txns,
        course_map=course_map, all_courses=all_courses)


@courses_bp.route('/admin/materials/<material_id>/transactions', methods=['POST'])
@admin_required
def admin_material_transaction(material_id):
    """新增進貨 / 銷售 / 調整記錄"""
    data = request.get_json() or {}
    txn_type = data.get('type')
    if txn_type not in ('in', 'out', 'adjust'):
        return jsonify({'error': '異動類型錯誤'}), 400

    qty = int(data.get('quantity') or 0)
    if qty <= 0:
        return jsonify({'error': '數量必須大於 0'}), 400

    # 銷售時確認庫存充足
    if txn_type == 'out':
        stock_row = supabase.table('material_stock').select('stock').eq('id', material_id).execute()
        current = (stock_row.data[0]['stock'] if stock_row.data else 0) or 0
        if qty > current:
            return jsonify({'error': f'庫存不足（現有 {current} {data.get("unit","本")}）'}), 400

    unit_cost = int(data.get('unit_cost') or 0)
    row = {
        'material_id': material_id,
        'type': txn_type,
        'quantity': qty,
        'unit_cost': unit_cost,
        'total_cost': unit_cost * qty if txn_type == 'in' else 0,
        'course_id': data.get('course_id') or None,
        'notes': (data.get('notes') or '').strip() or None,
        'transaction_date': data.get('transaction_date') or datetime.now(TAIPEI_TZ).date().isoformat(),
    }
    supabase.table('material_transactions').insert(row).execute()
    return jsonify({'success': True})
